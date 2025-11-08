import discord
from redbot.core import commands, Config, checks
from redbot.core.bot import Red
from datetime import datetime, timedelta
import asyncio
from typing import Optional, List, Dict, Tuple
from collections import defaultdict
import json
import logging
import os
import tempfile

log = logging.getLogger("red.chatsummary")


class ChatSummary(commands.Cog):
    """聊天频道总结插件
    
    支持总结指定频道和全部频道，可配置定时任务自动总结。
    """
    
    def __init__(self, bot: Red):
        self.bot = bot
        self.config = Config.get_conf(self, identifier=1234567890, force_registration=True)
        
        # 默认配置
        default_guild = {
            "enabled": False,
            "api_key": None,
            "api_base": "https://api.openai.com/v1",
            "model": "gpt-3.5-turbo",
            "max_messages": 100,  # 总结功能的最大消息数
            "export_max_messages": 1000,  # 导出功能的最大消息数
            "summary_channel": None,  # 总结消息发送的频道
            "export_channel": None,  # Excel导出文件发送的频道
            "scheduled_tasks": {},  # {channel_id: {"interval": hours, "enabled": True}}
            "export_tasks": {},  # {task_id: {"type": "all/category/channel", "target": "", "interval": hours, "enabled": True, "single_file": True}}
            "excluded_channels": [],  # 总结功能排除的频道
            "excluded_categories": [],  # 总结功能排除的分类列表
            "export_excluded_channels": [],  # 导出功能排除的频道
            "export_excluded_categories": [],  # 导出功能排除的分类列表
            "include_bots": False,
        }
        
        self.config.register_guild(**default_guild)
        
        # 定时任务字典
        self.scheduled_jobs = {}
        self.export_jobs = {}  # Excel导出定时任务
        
        # 启动时加载定时任务
        self.bot.loop.create_task(self.load_scheduled_tasks())
        self.bot.loop.create_task(self.load_export_tasks())
    
    def cog_unload(self):
        """卸载时取消所有定时任务"""
        for task in self.scheduled_jobs.values():
            task.cancel()
        for task in self.export_jobs.values():
            task.cancel()
    
    async def load_scheduled_tasks(self):
        """加载并启动所有已配置的定时任务"""
        await self.bot.wait_until_ready()
        
        for guild in self.bot.guilds:
            scheduled_tasks = await self.config.guild(guild).scheduled_tasks()
            
            for channel_id_str, task_config in scheduled_tasks.items():
                if task_config.get("enabled", False):
                    channel_id = int(channel_id_str)
                    interval = task_config.get("interval", 24)
                    self.start_scheduled_task(guild.id, channel_id, interval)
    
    async def load_export_tasks(self):
        """加载并启动所有已配置的导出定时任务"""
        await self.bot.wait_until_ready()
        
        for guild in self.bot.guilds:
            export_tasks = await self.config.guild(guild).export_tasks()
            
            for task_id, task_config in export_tasks.items():
                if task_config.get("enabled", False):
                    self.start_export_task(guild.id, task_id, task_config)
    
    def start_scheduled_task(self, guild_id: int, channel_id: int, interval_hours: int):
        """启动一个定时任务"""
        task_key = f"{guild_id}_{channel_id}"
        
        # 如果任务已存在，先取消
        if task_key in self.scheduled_jobs:
            self.scheduled_jobs[task_key].cancel()
        
        # 创建新任务
        task = self.bot.loop.create_task(
            self._scheduled_summary_loop(guild_id, channel_id, interval_hours)
        )
        self.scheduled_jobs[task_key] = task
    
    async def _scheduled_summary_loop(self, guild_id: int, channel_id: int, interval_hours: int):
        """定时任务循环"""
        while True:
            try:
                # 先等待指定时间
                await asyncio.sleep(interval_hours * 3600)
                
                guild = self.bot.get_guild(guild_id)
                if not guild:
                    continue
                
                # 检查是否是全服务器总结任务（channel_id 为 0）
                if channel_id == 0:
                    # 执行全服务器总结
                    await self._execute_all_summary(guild)
                else:
                    # 执行单个频道总结
                    channel = guild.get_channel(channel_id)
                    if not channel:
                        continue
                    await self._execute_summary(guild, channel)
                    
            except asyncio.CancelledError:
                break
            except Exception as e:
                log.error(f"定时任务执行错误 (Guild: {guild_id}, Channel: {channel_id}): {e}", exc_info=True)
    
    def start_export_task(self, guild_id: int, task_id: str, task_config: dict):
        """启动一个导出定时任务"""
        task_key = f"{guild_id}_{task_id}"
        
        # 如果任务已存在，先取消
        if task_key in self.export_jobs:
            self.export_jobs[task_key].cancel()
        
        # 创建新任务
        task = self.bot.loop.create_task(
            self._export_task_loop(guild_id, task_id, task_config)
        )
        self.export_jobs[task_key] = task
    
    async def _export_task_loop(self, guild_id: int, task_id: str, task_config: dict):
        """导出定时任务循环"""
        while True:
            try:
                # 先等待指定时间
                interval_hours = task_config.get("interval", 24)
                await asyncio.sleep(interval_hours * 3600)
                
                guild = self.bot.get_guild(guild_id)
                if not guild:
                    continue
                
                # 执行导出任务
                await self._execute_export_task(guild, task_config)
                    
            except asyncio.CancelledError:
                break
            except Exception as e:
                log.error(f"导出定时任务执行错误 (Guild: {guild_id}, Task: {task_id}): {e}", exc_info=True)
    
    async def _execute_export_task(self, guild: discord.Guild, task_config: dict):
        """执行导出任务"""
        try:
            task_type = task_config.get("type", "all")
            target = task_config.get("target", "")
            single_file = task_config.get("single_file", True)
            max_messages = task_config.get("max_messages", 0)
            
            # 获取发送目标频道（优先使用导出频道，否则使用总结频道）
            export_channel_id = await self.config.guild(guild).export_channel()
            summary_channel_id = await self.config.guild(guild).summary_channel()
            
            target_channel = None
            if export_channel_id:
                target_channel = guild.get_channel(export_channel_id)
            elif summary_channel_id:
                target_channel = guild.get_channel(summary_channel_id)
            
            if not target_channel:
                target_channel = guild.system_channel or guild.text_channels[0] if guild.text_channels else None
            
            if not target_channel:
                log.error(f"无法找到发送导出文件的频道 (Guild: {guild.name})")
                return
            
            max_msgs = max_messages if max_messages > 0 else None
            
            if task_type == "all":
                # 导出所有频道
                log.info(f"执行定时导出任务：所有频道 (Guild: {guild.name}, 单文件: {single_file})")
                
                categories_dict = defaultdict(list)
                for channel in guild.text_channels:
                    if await self._is_channel_excluded_from_export(guild, channel):
                        continue
                    category_name = channel.category.name if channel.category else "未分类"
                    categories_dict[category_name].append(channel)
                
                if not categories_dict:
                    log.warning(f"没有可导出的频道 (Guild: {guild.name})")
                    return
                
                if single_file:
                    # 单文件模式
                    await target_channel.send("📄 正在执行定时导出任务（所有频道）...")
                    report_title = "全服务器聊天记录"
                    excel_path = await self.generate_multi_channel_excel_report(
                        guild, categories_dict, report_title, max_msgs
                    )
                    
                    if excel_path and os.path.exists(excel_path):
                        file_size = os.path.getsize(excel_path) / (1024 * 1024)
                        await target_channel.send(
                            f"✅ 定时导出完成（文件大小: {file_size:.2f} MB）",
                            file=discord.File(excel_path, filename=f"{guild.name}_全服务器聊天记录.xlsx")
                        )
                        os.remove(excel_path)
                        log.info(f"成功完成定时导出任务：所有频道 (Guild: {guild.name})")
                else:
                    # 多文件模式
                    await target_channel.send("📄 正在执行定时导出任务（所有频道，多文件模式）...")
                    total_exported = 0
                    for category_name in sorted(categories_dict.keys(), key=lambda x: (x == "未分类", x)):
                        for channel in sorted(categories_dict[category_name], key=lambda c: c.position):
                            try:
                                excel_path = await self.generate_excel_report(channel, max_msgs)
                                if excel_path and os.path.exists(excel_path):
                                    await target_channel.send(
                                        file=discord.File(excel_path, filename=f"{category_name}-{channel.name}.xlsx")
                                    )
                                    os.remove(excel_path)
                                    total_exported += 1
                                await asyncio.sleep(2)
                            except Exception as e:
                                log.error(f"导出频道 {channel.name} 时出错: {e}")
                    await target_channel.send(f"✅ 定时导出完成！共导出 {total_exported} 个频道。")
                    
            elif task_type == "category":
                # 导出指定分类
                log.info(f"执行定时导出任务：分类 {target} (Guild: {guild.name}, 单文件: {single_file})")
                
                channels_in_category = []
                if target == "未分类":
                    for channel in guild.text_channels:
                        if not channel.category and not await self._is_channel_excluded_from_export(guild, channel):
                            channels_in_category.append(channel)
                else:
                    for category in guild.categories:
                        if category.name == target:
                            for channel in category.text_channels:
                                if not await self._is_channel_excluded_from_export(guild, channel):
                                    channels_in_category.append(channel)
                            break
                
                if not channels_in_category:
                    log.warning(f"分类 {target} 中没有可导出的频道 (Guild: {guild.name})")
                    return
                
                if single_file:
                    await target_channel.send(f"📄 正在执行定时导出任务（分类：{target}）...")
                    categories_dict = {target: channels_in_category}
                    report_title = f"{target}_聊天记录"
                    excel_path = await self.generate_multi_channel_excel_report(
                        guild, categories_dict, report_title, max_msgs
                    )
                    
                    if excel_path and os.path.exists(excel_path):
                        file_size = os.path.getsize(excel_path) / (1024 * 1024)
                        await target_channel.send(
                            f"✅ 定时导出完成（分类：{target}，文件大小: {file_size:.2f} MB）",
                            file=discord.File(excel_path, filename=f"{target}_聊天记录.xlsx")
                        )
                        os.remove(excel_path)
                        log.info(f"成功完成定时导出任务：分类 {target} (Guild: {guild.name})")
                else:
                    await target_channel.send(f"📄 正在执行定时导出任务（分类：{target}，多文件模式）...")
                    total_exported = 0
                    for channel in sorted(channels_in_category, key=lambda c: c.position):
                        try:
                            excel_path = await self.generate_excel_report(channel, max_msgs)
                            if excel_path and os.path.exists(excel_path):
                                await target_channel.send(
                                    file=discord.File(excel_path, filename=f"{target}-{channel.name}.xlsx")
                                )
                                os.remove(excel_path)
                                total_exported += 1
                            await asyncio.sleep(2)
                        except Exception as e:
                            log.error(f"导出频道 {channel.name} 时出错: {e}")
                    await target_channel.send(f"✅ 定时导出完成（分类：{target}）！共导出 {total_exported} 个频道。")
                    
            elif task_type == "channel":
                # 导出指定频道
                channel_id = int(target)
                channel = guild.get_channel(channel_id)
                
                if not channel:
                    log.error(f"找不到频道 {channel_id} (Guild: {guild.name})")
                    return
                
                log.info(f"执行定时导出任务：频道 {channel.name} (Guild: {guild.name})")
                await target_channel.send(f"📄 正在执行定时导出任务（频道：{channel.mention}）...")
                
                excel_path = await self.generate_excel_report(channel, max_msgs)
                if excel_path and os.path.exists(excel_path):
                    category_name = channel.category.name if channel.category else "未分类"
                    await target_channel.send(
                        f"✅ 定时导出完成（频道：{channel.mention}）",
                        file=discord.File(excel_path, filename=f"{category_name}-{channel.name}.xlsx")
                    )
                    os.remove(excel_path)
                    log.info(f"成功完成定时导出任务：频道 {channel.name} (Guild: {guild.name})")
                    
        except Exception as e:
            log.error(f"执行导出任务时出错 (Guild: {guild.name}): {e}", exc_info=True)
    
    async def _execute_summary(self, guild: discord.Guild, channel: discord.TextChannel):
        """执行单个频道总结并发送结果"""
        try:
            # 生成总结
            summary = await self.generate_channel_summary(channel)
            
            # 发送到指定频道
            summary_channel_id = await self.config.guild(guild).summary_channel()
            if summary_channel_id:
                summary_channel = guild.get_channel(summary_channel_id)
                if summary_channel:
                    await summary_channel.send(embed=summary)
                    log.info(f"总结已发送到指定频道 {summary_channel.name} (Guild: {guild.name})")
                else:
                    log.warning(f"配置的总结频道不存在 (ID: {summary_channel_id}, Guild: {guild.name})")
                    await channel.send(embed=summary)
            else:
                # 发送到原频道
                await channel.send(embed=summary)
                log.info(f"总结已发送到原频道 {channel.name} (Guild: {guild.name})")
        except Exception as e:
            log.error(f"执行总结时出错 (Channel: {channel.name}, Guild: {guild.name}): {e}", exc_info=True)
    
    async def _is_channel_excluded(self, guild: discord.Guild, channel: discord.TextChannel) -> bool:
        """检查频道是否应该被排除（基于频道本身或其分类）- 用于总结功能"""
        excluded_channels = await self.config.guild(guild).excluded_channels()
        excluded_categories = await self.config.guild(guild).excluded_categories()
        
        # 检查频道是否被排除
        if channel.id in excluded_channels:
            return True
        
        # 检查分类是否被排除
        if channel.category and channel.category.name in excluded_categories:
            return True
        
        # 检查未分类频道（如果"未分类"在排除列表中）
        if not channel.category and "未分类" in excluded_categories:
            return True
        
        return False
    
    async def _is_channel_excluded_from_export(self, guild: discord.Guild, channel: discord.TextChannel) -> bool:
        """检查频道是否应该被排除（基于频道本身或其分类）- 用于导出功能"""
        excluded_channels = await self.config.guild(guild).export_excluded_channels()
        excluded_categories = await self.config.guild(guild).export_excluded_categories()
        
        # 检查频道是否被排除
        if channel.id in excluded_channels:
            return True
        
        # 检查分类是否被排除
        if channel.category and channel.category.name in excluded_categories:
            return True
        
        # 检查未分类频道（如果"未分类"在排除列表中）
        if not channel.category and "未分类" in excluded_categories:
            return True
        
        return False
    
    async def _execute_all_summary(self, guild: discord.Guild):
        """执行全服务器总结并发送结果（包括PDF生成）"""
        try:
            # 按分类分组频道
            categories_dict = defaultdict(list)
            
            for channel in guild.text_channels:
                # 使用新的检查方法
                if await self._is_channel_excluded(guild, channel):
                    continue
                
                category_name = channel.category.name if channel.category else "未分类"
                categories_dict[category_name].append(channel)
            
            if not categories_dict:
                log.warning(f"没有可总结的频道 (Guild: {guild.name})")
                return
            
            # 获取发送目标频道
            summary_channel_id = await self.config.guild(guild).summary_channel()
            target_channel = guild.get_channel(summary_channel_id) if summary_channel_id else None
            
            if not target_channel:
                # 如果没有配置总结频道，尝试找一个默认频道
                target_channel = guild.system_channel or guild.text_channels[0] if guild.text_channels else None
            
            if not target_channel:
                log.error(f"无法找到发送总结的频道 (Guild: {guild.name})")
                return
            
            # 发送报告标题
            await target_channel.send(f"## 📊 服务器全频道总结报告\n生成时间: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC")
            
            total_channels = 0
            summaries_data = []  # 收集PDF数据
            
            # 按分类名称排序（"未分类"放在最后）
            sorted_categories = sorted(categories_dict.keys(), key=lambda x: (x == "未分类", x))
            
            for category_name in sorted_categories:
                channels = categories_dict[category_name]
                
                # 发送分类标题
                await target_channel.send(f"\n## 📁 {category_name}\n")
                
                # 总结该分类下的所有频道
                for channel in sorted(channels, key=lambda c: c.position):
                    try:
                        summary_embed = await self.generate_channel_summary(channel)
                        await target_channel.send(embed=summary_embed)
                        total_channels += 1
                        
                        # 收集PDF数据
                        summary_text = summary_embed.description or "无总结内容"
                        stats = {}
                        for field in summary_embed.fields:
                            if "消息数量" in field.name:
                                stats['message_count'] = field.value
                            elif "参与人数" in field.name:
                                stats['user_count'] = field.value
                            elif "时间范围" in field.name:
                                stats['time_range'] = field.value
                        
                        summaries_data.append({
                            'category': category_name,
                            'channel_name': channel.name,
                            'summary_text': summary_text,
                            'stats': stats
                        })
                        
                        log.info(f"成功总结频道 {channel.name} (分类: {category_name}, Guild: {guild.name})")
                        await asyncio.sleep(1)  # 避免速率限制
                    except Exception as e:
                        log.error(f"总结频道 {channel.name} 时出错 (分类: {category_name}, Guild: {guild.name}): {e}", exc_info=True)
            
            # 发送完成消息
            await target_channel.send(f"✅ 定时总结完成！共总结了 {total_channels} 个频道，分布在 {len(categories_dict)} 个分类中。")
            log.info(f"完成全服务器总结 (Guild: {guild.name}, 总频道数: {total_channels})")
            
            # 生成并发送PDF
            if summaries_data:
                await target_channel.send("📄 正在生成PDF报告...")
                report_title = f"{guild.name} - Server Summary Report"
                pdf_path = await self.generate_pdf_report(guild, summaries_data, report_title)
                
                if pdf_path and os.path.exists(pdf_path):
                    try:
                        await target_channel.send(
                            "📊 总结报告PDF文件：",
                            file=discord.File(pdf_path, filename=f"summary_{guild.name}_{datetime.utcnow().strftime('%Y%m%d')}.pdf")
                        )
                        log.info(f"成功发送PDF报告 (Guild: {guild.name})")
                        # 删除临时文件
                        os.remove(pdf_path)
                    except Exception as e:
                        log.error(f"发送PDF文件时出错: {e}", exc_info=True)
                        await target_channel.send("❌ PDF文件生成成功但发送失败。")
                else:
                    await target_channel.send("❌ PDF文件生成失败。请检查日志。")
            
        except Exception as e:
            log.error(f"执行全服务器总结时出错 (Guild: {guild.name}): {e}", exc_info=True)
    
    async def generate_channel_summary(self, channel: discord.TextChannel) -> discord.Embed:
        """生成频道总结"""
        guild = channel.guild
        max_messages = await self.config.guild(guild).max_messages()
        include_bots = await self.config.guild(guild).include_bots()
        
        # 获取消息
        messages = []
        async for message in channel.history(limit=max_messages):
            if not include_bots and message.author.bot:
                continue
            messages.append(message)
        
        messages.reverse()  # 按时间顺序排列
        
        # 获取频道分类
        category_name = channel.category.name if channel.category else "未分类"
        
        if not messages:
            embed = discord.Embed(
                title=f"📊 频道总结 - {category_name} / {channel.name}",
                description="没有找到消息记录。",
                color=discord.Color.blue(),
                timestamp=datetime.utcnow()
            )
            return embed
        
        # 生成总结
        summary_text = await self.summarize_messages(guild, messages)
        
        # 创建统计信息
        user_count = len(set(m.author.id for m in messages))
        time_range = f"{messages[0].created_at.strftime('%Y-%m-%d %H:%M')} - {messages[-1].created_at.strftime('%Y-%m-%d %H:%M')}"
        
        embed = discord.Embed(
            title=f"📊 频道总结 - {category_name} / {channel.name}",
            description=summary_text,
            color=discord.Color.green(),
            timestamp=datetime.utcnow()
        )
        
        embed.add_field(name="📝 消息数量", value=str(len(messages)), inline=True)
        embed.add_field(name="👥 参与人数", value=str(user_count), inline=True)
        embed.add_field(name="⏰ 时间范围", value=time_range, inline=False)
        
        return embed
    
    async def summarize_messages(self, guild: discord.Guild, messages: List[discord.Message]) -> str:
        """使用 AI 总结消息"""
        api_key = await self.config.guild(guild).api_key()
        
        if not api_key:
            # 如果没有配置 API key，使用简单统计
            return self.simple_summary(messages)
        
        # 准备消息文本
        message_text = "\n".join([
            f"[{msg.created_at.strftime('%H:%M')}] {msg.author.name}: {msg.content[:200]}"
            for msg in messages
            if msg.content
        ])
        
        if not message_text:
            return "没有文本消息可以总结。"
        
        # 调用 AI API
        try:
            import aiohttp
            
            api_base = await self.config.guild(guild).api_base()
            model = await self.config.guild(guild).model()
            
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            
            prompt = f"""You are an **expert in summarizing Discord content**, skilled at extracting key information and generating **high-quality, well-structured summaries**.
Based on the provided Video Transcript, complete the following tasks:

**Task Description:**
Act as a helpful assistant. Your task is to summarize the key points from [meeting notes]. The summary should be concise yet comprehensive, capturing the essence of the meeting. Your summary should enable someone who wasn't present at the meeting to understand its outcomes and next steps clearly.Length: - Ensure the response has a minimum of 800 words

Language: - The entire output, including **section titles and labels**, must be written in the "简体中文" language (For example, Summary, Highlights, Key Insights, Outline, Core Concepts, Keywords, FAQ, etc. all need to be translated into 简体中文 language.).
- Do **not** include any separators (`---`), or additional text outside of the task results.

The Discord content:
{message_text[:4000]}"""
            
            data = {
                "model": model,
                "messages": [
                    {"role": "system", "content": "你是一个专业的聊天记录总结助手。"},
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 500,
                "temperature": 0.7
            }
            
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    f"{api_base}/chat/completions",
                    headers=headers,
                    json=data,
                    timeout=aiohttp.ClientTimeout(total=30)
                ) as resp:
                    if resp.status == 200:
                        result = await resp.json()
                        return result["choices"][0]["message"]["content"]
                    else:
                        return f"API 调用失败（状态码: {resp.status}），使用简单统计。\n\n" + self.simple_summary(messages)
        
        except Exception as e:
            return f"总结生成失败: {str(e)}\n\n使用简单统计:\n{self.simple_summary(messages)}"
    
    def simple_summary(self, messages: List[discord.Message]) -> str:
        """简单的统计总结（不使用 AI）"""
        if not messages:
            return "没有消息记录。"
        
        # 统计活跃用户
        user_msg_count = {}
        for msg in messages:
            user_msg_count[msg.author.name] = user_msg_count.get(msg.author.name, 0) + 1
        
        top_users = sorted(user_msg_count.items(), key=lambda x: x[1], reverse=True)[:5]
        
        summary = "**活跃用户统计：**\n"
        for i, (user, count) in enumerate(top_users, 1):
            summary += f"{i}. {user}: {count} 条消息\n"
        
        return summary
    
    def _parse_markdown_to_pdf_elements(self, markdown_text: str, styles: dict, use_chinese: bool):
        """将Markdown文本解析为PDF元素列表
        
        参数:
            markdown_text: Markdown格式的文本
            styles: PDF样式字典
            use_chinese: 是否使用中文字体
        
        返回:
            PDF元素列表（Paragraph和Spacer对象）
        """
        from reportlab.platypus import Paragraph, Spacer
        from reportlab.lib.units import cm
        import re
        
        elements = []
        lines = markdown_text.split('\n')
        i = 0
        
        # 选择样式
        h1_style = styles.get('ChineseH1', styles['Heading1']) if use_chinese else styles['Heading1']
        h2_style = styles.get('ChineseH2', styles['Heading2']) if use_chinese else styles['Heading2']
        h3_style = styles.get('ChineseH3', styles['Heading3']) if use_chinese else styles['Heading3']
        body_style = styles.get('ChineseBody', styles['BodyText']) if use_chinese else styles['BodyText']
        
        max_iterations = len(lines) * 2  # 防止无限循环
        iteration_count = 0
        
        while i < len(lines):
            iteration_count += 1
            if iteration_count > max_iterations:
                log.error(f"Markdown解析超过最大迭代次数，强制退出。当前行: {i}/{len(lines)}")
                break
            
            line = lines[i].strip()
            
            # 跳过空行
            if not line:
                i += 1
                continue
            
            # 清理XML特殊字符
            def clean_text(text):
                return (text.replace('&', '&amp;')
                           .replace('<', '&lt;')
                           .replace('>', '&gt;'))
            
            # 解析一级标题 # 标题
            if line.startswith('# ') and not line.startswith('## '):
                title_text = clean_text(line[2:])
                elements.append(Paragraph(f"<b>{title_text}</b>", h1_style))
                elements.append(Spacer(1, 0.3*cm))
                i += 1
            
            # 解析二级标题 ## 标题
            elif line.startswith('## ') and not line.startswith('### '):
                title_text = clean_text(line[3:])
                elements.append(Paragraph(f"<b>{title_text}</b>", h2_style))
                elements.append(Spacer(1, 0.2*cm))
                i += 1
            
            # 解析三级标题 ### 标题
            elif line.startswith('### '):
                title_text = clean_text(line[4:])
                elements.append(Paragraph(f"<b>{title_text}</b>", h3_style))
                elements.append(Spacer(1, 0.2*cm))
                i += 1
            
            # 解析列表项 - 或 * 开头
            elif line.startswith('- ') or line.startswith('* '):
                # 收集连续的列表项
                list_items = []
                while i < len(lines):
                    current_line = lines[i].strip()
                    if current_line.startswith('- ') or current_line.startswith('* '):
                        item_text = clean_text(current_line[2:])
                        # 处理列表项中的粗体标记
                        item_text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', item_text)
                        list_items.append(item_text)
                        i += 1
                    elif not current_line:
                        i += 1
                        break
                    else:
                        break
                
                # 生成列表
                for item in list_items:
                    elements.append(Paragraph(f"• {item}", body_style))
                elements.append(Spacer(1, 0.2*cm))
            
            # 解析粗体 **文本**
            elif '**' in line:
                # 处理粗体标记
                text = clean_text(line)
                text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)
                elements.append(Paragraph(text, body_style))
                elements.append(Spacer(1, 0.15*cm))
                i += 1
            
            # 普通段落
            else:
                # 收集连续的非空行作为一个段落
                para_lines = []
                while i < len(lines):
                    current_line = lines[i].strip()
                    if current_line and not current_line.startswith('#') and not current_line.startswith('-') and not current_line.startswith('*'):
                        para_lines.append(current_line)
                        i += 1
                    else:
                        break
                
                if para_lines:
                    para_text = clean_text(' '.join(para_lines))
                    # 处理粗体
                    para_text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', para_text)
                    elements.append(Paragraph(para_text, body_style))
                    elements.append(Spacer(1, 0.2*cm))
                else:
                    # 如果没有收集到任何段落行，说明遇到了无法处理的行
                    # 强制跳过以避免无限循环
                    i += 1
        
        return elements
    
    async def generate_pdf_report(self, guild: discord.Guild, summaries_data: List[Dict], report_title: str) -> str:
        """生成PDF报告（异步包装器）
        
        在单独的线程中运行PDF生成，避免阻塞Discord事件循环
        
        参数:
            guild: Discord服务器
            summaries_data: 总结数据列表
            report_title: 报告标题
        
        返回:
            PDF文件路径
        """
        try:
            log.info(f"开始生成PDF报告 (Guild: {guild.name}, 频道数: {len(summaries_data)})")
            
            # 深拷贝数据以避免线程安全问题
            import copy
            summaries_data_copy = copy.deepcopy(summaries_data)
            
            # 在线程池中运行同步PDF生成函数
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                None,
                self._generate_pdf_report_sync,
                guild.name,
                guild.id,
                summaries_data_copy,
                report_title
            )
            
            log.info(f"PDF报告生成完成: {result}")
            return result
            
        except Exception as e:
            log.error(f"异步PDF生成包装器出错: {e}", exc_info=True)
            return None
    
    def _generate_pdf_report_sync(self, guild_name: str, guild_id: int, summaries_data: List[Dict], report_title: str) -> str:
        """生成PDF报告（同步版本，在单独线程中运行）
        
        参数:
            guild_name: Discord服务器名称
            guild_id: Discord服务器ID
            summaries_data: 总结数据列表，每项包含 category, channel_name, summary_text, stats
            report_title: 报告标题
        
        返回:
            PDF文件路径
        """
        try:
            log.info(f"[PDF线程] 开始同步生成PDF (频道数: {len(summaries_data)})")
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Flowable
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.enums import TA_LEFT, TA_CENTER
            
            # 创建书签Flowable
            class BookmarkFlowable(Flowable):
                """在PDF中添加书签的Flowable"""
                def __init__(self, title, key):
                    Flowable.__init__(self)
                    self.title = title
                    self.key = key
                    self.width = 0
                    self.height = 0
                
                def draw(self):
                    """在当前位置添加书签"""
                    self.canv.bookmarkPage(self.key)
                    self.canv.addOutlineEntry(self.title, self.key, level=0)
            
            # 创建临时文件
            temp_dir = tempfile.gettempdir()
            pdf_filename = f"summary_{guild_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            pdf_path = os.path.join(temp_dir, pdf_filename)
            
            # 创建PDF文档
            doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                                   rightMargin=2*cm, leftMargin=2*cm,
                                   topMargin=2*cm, bottomMargin=2*cm)
            
            story = []
            styles = getSampleStyleSheet()
            
            # 尝试注册中文字体（如果失败则使用默认字体）
            try:
                # 常见的中文字体路径（包括TTC和TTF文件）
                chinese_fonts = [
                    # macOS
                    ('/System/Library/Fonts/PingFang.ttc', 0),
                    ('/System/Library/Fonts/STHeiti Light.ttc', 0),
                    ('/System/Library/Fonts/Hiragino Sans GB.ttc', 0),
                    # Linux
                    ('/usr/share/fonts/truetype/arphic/uming.ttc', 0),
                    ('/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf', None),
                    ('/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc', 0),
                    # Windows
                    ('C:\\Windows\\Fonts\\msyh.ttc', 0),
                    ('C:\\Windows\\Fonts\\simhei.ttf', None),
                    ('C:\\Windows\\Fonts\\simsun.ttc', 0),
                ]
                
                font_registered = False
                for font_path, subfont_index in chinese_fonts:
                    if os.path.exists(font_path):
                        try:
                            # TTC文件需要指定subfontIndex
                            if subfont_index is not None:
                                pdfmetrics.registerFont(TTFont('Chinese', font_path, subfontIndex=subfont_index))
                            else:
                                pdfmetrics.registerFont(TTFont('Chinese', font_path))
                            font_registered = True
                            log.info(f"成功注册中文字体: {font_path}")
                            break
                        except Exception as e:
                            log.debug(f"尝试注册字体 {font_path} 失败: {e}")
                            continue
                
                if font_registered:
                    # 创建中文样式
                    styles.add(ParagraphStyle(name='ChineseTitle',
                                             parent=styles['Heading1'],
                                             fontName='Chinese',
                                             fontSize=18,
                                             alignment=TA_CENTER,
                                             wordWrap='CJK'))
                    styles.add(ParagraphStyle(name='ChineseHeading',
                                             parent=styles['Heading2'],
                                             fontName='Chinese',
                                             fontSize=14,
                                             wordWrap='CJK'))
                    styles.add(ParagraphStyle(name='ChineseH1',
                                             parent=styles['Heading1'],
                                             fontName='Chinese',
                                             fontSize=16,
                                             wordWrap='CJK',
                                             spaceAfter=12))
                    styles.add(ParagraphStyle(name='ChineseH2',
                                             parent=styles['Heading2'],
                                             fontName='Chinese',
                                             fontSize=13,
                                             wordWrap='CJK',
                                             spaceAfter=10))
                    styles.add(ParagraphStyle(name='ChineseH3',
                                             parent=styles['Heading3'],
                                             fontName='Chinese',
                                             fontSize=11,
                                             wordWrap='CJK',
                                             spaceAfter=8))
                    styles.add(ParagraphStyle(name='ChineseBody',
                                             parent=styles['BodyText'],
                                             fontName='Chinese',
                                             fontSize=10,
                                             wordWrap='CJK',
                                             leading=14))
                    use_chinese = True
                else:
                    log.warning("未找到可用的中文字体，PDF将使用默认字体（中文可能显示为方块）")
                    use_chinese = False
            except Exception as e:
                log.error(f"注册中文字体时出错: {e}", exc_info=True)
                use_chinese = False
            
            # 选择样式
            title_style = styles['ChineseTitle'] if use_chinese else styles['Title']
            heading_style = styles['ChineseHeading'] if use_chinese else styles['Heading2']
            body_style = styles['ChineseBody'] if use_chinese else styles['BodyText']
            
            # 添加标题
            story.append(Paragraph(report_title, title_style))
            story.append(Spacer(1, 0.5*cm))
            
            # 添加生成时间
            gen_time = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
            story.append(Paragraph(gen_time, body_style))
            story.append(Spacer(1, 0.3*cm))
            
            # 添加服务器信息
            server_info = f"Server: {guild_name}"
            story.append(Paragraph(server_info, body_style))
            story.append(Spacer(1, 1*cm))
            
            # 添加每个频道的总结
            log.info(f"[PDF线程] 开始处理 {len(summaries_data)} 个频道的内容")
            for i, data in enumerate(summaries_data):
                try:
                    # 分类和频道标题
                    log.info(f"[PDF线程] 正在处理频道 {i+1}/{len(summaries_data)}")
                    category = data.get('category', '未知分类')
                    channel_name = data.get('channel_name', '未知频道')
                    title = f"{category} / {channel_name}"
                    
                    log.info(f"[PDF线程] 频道标题: {title}")
                    
                    # 添加书签（使用BookmarkFlowable）
                    log.info(f"[PDF线程] 添加书签")
                    bookmark_key = f"channel_{i}"
                    story.append(BookmarkFlowable(title, bookmark_key))
                    
                    # 添加频道标题
                    log.info(f"[PDF线程] 添加标题段落")
                    story.append(Paragraph(title, heading_style))
                    story.append(Spacer(1, 0.3*cm))
                    
                    # 总结内容 - 使用Markdown解析器
                    summary_text = data.get('summary_text', '无总结内容')
                    log.info(f"[PDF线程] 开始解析Markdown，文本长度: {len(summary_text)}")
                    
                    # 解析Markdown并添加到story
                    try:
                        # 添加超时保护
                        import signal
                        
                        def timeout_handler(signum, frame):
                            raise TimeoutError("Markdown解析超时")
                        
                        # 设置10秒超时（仅在Unix系统上有效）
                        old_handler = None
                        try:
                            old_handler = signal.signal(signal.SIGALRM, timeout_handler)
                            signal.alarm(10)
                        except (AttributeError, ValueError):
                            # Windows系统不支持SIGALRM，跳过超时设置
                            pass
                        
                        try:
                            markdown_elements = self._parse_markdown_to_pdf_elements(summary_text, styles, use_chinese)
                            log.info(f"[PDF线程] Markdown解析完成，生成 {len(markdown_elements)} 个元素")
                            story.extend(markdown_elements)
                        finally:
                            # 取消超时
                            try:
                                signal.alarm(0)
                                if old_handler:
                                    signal.signal(signal.SIGALRM, old_handler)
                            except (AttributeError, ValueError):
                                pass
                        
                    except TimeoutError:
                        log.error(f"[PDF线程] Markdown解析超时 ({title})，使用简单文本")
                        # 使用简单文本作为备选
                        clean_text = summary_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        story.append(Paragraph(clean_text[:500] + "...", body_style))
                    except Exception as e:
                        log.error(f"[PDF线程] 解析Markdown时出错 ({title}): {e}", exc_info=True)
                        # 使用简单文本作为备选
                        clean_text = summary_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        story.append(Paragraph(clean_text[:500] + "...", body_style))
                    
                except Exception as e:
                    log.error(f"[PDF线程] 处理频道 {i+1} 时出错: {e}", exc_info=True)
                    continue
                
                # 统计信息
                stats = data.get('stats', {})
                stats_text = f"Messages: {stats.get('message_count', 0)} | Users: {stats.get('user_count', 0)} | Time: {stats.get('time_range', 'N/A')}"
                story.append(Paragraph(stats_text, body_style))
                story.append(Spacer(1, 0.3*cm))
                
                # 如果不是最后一个，添加分页
                if i < len(summaries_data) - 1:
                    story.append(PageBreak())
            
            # 生成PDF
            log.info(f"[PDF线程] 开始构建PDF文档，总元素数: {len(story)}")
            doc.build(story)
            log.info(f"[PDF线程] 成功生成PDF报告（包含 {len(summaries_data)} 个书签）: {pdf_path}")
            return pdf_path
            
        except ImportError as e:
            log.error(f"[PDF线程] reportlab库未安装，无法生成PDF: {e}")
            return None
        except Exception as e:
            log.error(f"[PDF线程] 生成PDF时出错: {e}", exc_info=True)
            return None
    
    @commands.group(name="summary", aliases=["总结"])
    @commands.guild_only()
    async def summary(self, ctx: commands.Context):
        """聊天总结命令组"""
        if ctx.invoked_subcommand is None:
            await ctx.send_help(ctx.command)
    
    @summary.command(name="channel", aliases=["频道"])
    async def summary_channel(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None):
        """总结指定频道的聊天记录
        
        参数:
            channel: 要总结的频道（不指定则总结当前频道）
        """
        if not await self.config.guild(ctx.guild).enabled():
            await ctx.send("❌ 聊天总结功能未启用。请管理员使用 `[p]summary enable` 启用。")
            return
        
        target_channel = channel or ctx.channel
        
        async with ctx.typing():
            summary_embed = await self.generate_channel_summary(target_channel)
            await ctx.send(embed=summary_embed)
    
    @summary.command(name="all", aliases=["全部", "全部频道"])
    @checks.admin_or_permissions(manage_guild=True)
    async def summary_all(self, ctx: commands.Context, generate_pdf: bool = True):
        """总结服务器中所有文字频道（需要管理员权限）
        
        参数:
            generate_pdf: 是否生成PDF文件（默认为 True）
        """
        if not await self.config.guild(ctx.guild).enabled():
            await ctx.send("❌ 聊天总结功能未启用。")
            return
        
        await ctx.send("🔄 开始总结所有频道，这可能需要一些时间...")
        
        # 按分类分组频道
        categories_dict = defaultdict(list)
        
        for channel in ctx.guild.text_channels:
            # 使用新的检查方法
            if await self._is_channel_excluded(ctx.guild, channel):
                continue
            
            category_name = channel.category.name if channel.category else "未分类"
            categories_dict[category_name].append(channel)
        
        if not categories_dict:
            await ctx.send("❌ 没有可总结的频道。")
            return
        
        # 发送到指定频道或当前频道
        summary_channel_id = await self.config.guild(ctx.guild).summary_channel()
        target_channel = ctx.guild.get_channel(summary_channel_id) if summary_channel_id else ctx.channel
        
        await target_channel.send(f"## 📊 服务器全频道总结报告\n生成时间: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC")
        
        total_channels = 0
        summaries_data = []  # 收集PDF数据
        
        # 按分类名称排序（"未分类"放在最后）
        sorted_categories = sorted(categories_dict.keys(), key=lambda x: (x == "未分类", x))
        
        for category_name in sorted_categories:
            channels = categories_dict[category_name]
            
            # 发送分类标题
            await target_channel.send(f"\n## 📁 {category_name}\n")
            
            # 总结该分类下的所有频道
            for channel in sorted(channels, key=lambda c: c.position):
                try:
                    summary_embed = await self.generate_channel_summary(channel)
                    await target_channel.send(embed=summary_embed)
                    total_channels += 1
                    
                    # 收集PDF数据
                    if generate_pdf:
                        # 从embed提取数据
                        summary_text = summary_embed.description or "无总结内容"
                        stats = {}
                        for field in summary_embed.fields:
                            if "消息数量" in field.name:
                                stats['message_count'] = field.value
                            elif "参与人数" in field.name:
                                stats['user_count'] = field.value
                            elif "时间范围" in field.name:
                                stats['time_range'] = field.value
                        
                        summaries_data.append({
                            'category': category_name,
                            'channel_name': channel.name,
                            'summary_text': summary_text,
                            'stats': stats
                        })
                    
                    log.info(f"成功总结频道 {channel.name} (分类: {category_name}, Guild: {ctx.guild.name})")
                    await asyncio.sleep(1)  # 避免速率限制
                except Exception as e:
                    log.error(f"总结频道 {channel.name} 时出错 (分类: {category_name}, Guild: {ctx.guild.name}): {e}", exc_info=True)
        
        await ctx.send(f"✅ 总结完成！共总结了 {total_channels} 个频道，分布在 {len(categories_dict)} 个分类中。")
        
        # 生成并发送PDF
        if generate_pdf and summaries_data:
            await ctx.send("📄 正在生成PDF报告...")
            report_title = f"{ctx.guild.name} - Server Summary Report"
            pdf_path = await self.generate_pdf_report(ctx.guild, summaries_data, report_title)
            
            if pdf_path and os.path.exists(pdf_path):
                try:
                    await target_channel.send(
                        "📊 总结报告PDF文件：",
                        file=discord.File(pdf_path, filename=f"summary_{ctx.guild.name}_{datetime.utcnow().strftime('%Y%m%d')}.pdf")
                    )
                    log.info(f"成功发送PDF报告 (Guild: {ctx.guild.name})")
                    # 删除临时文件
                    os.remove(pdf_path)
                except Exception as e:
                    log.error(f"发送PDF文件时出错: {e}", exc_info=True)
                    await ctx.send("❌ PDF文件生成成功但发送失败。")
            else:
                await ctx.send("❌ PDF文件生成失败。请检查日志。")
    
    @summary.command(name="category", aliases=["分类"])
    @checks.admin_or_permissions(manage_guild=True)
    async def summary_category(self, ctx: commands.Context, category_name: str, generate_pdf: bool = True):
        """总结指定分类下的所有频道（需要管理员权限）
        
        参数:
            category_name: 分类名称（使用"未分类"总结没有分类的频道）
            generate_pdf: 是否生成PDF文件（默认为 True）
        """
        if not await self.config.guild(ctx.guild).enabled():
            await ctx.send("❌ 聊天总结功能未启用。")
            return
        
        # 查找分类下的频道
        channels_in_category = []
        
        if category_name == "未分类":
            # 收集所有未分类的频道
            for channel in ctx.guild.text_channels:
                if not channel.category:
                    # 检查是否被排除
                    if not await self._is_channel_excluded(ctx.guild, channel):
                        channels_in_category.append(channel)
        else:
            # 查找指定分类
            target_category = None
            for category in ctx.guild.categories:
                if category.name == category_name:
                    target_category = category
                    break
            
            if not target_category:
                await ctx.send(f"❌ 找不到名为 `{category_name}` 的分类。")
                return
            
            # 收集该分类下的所有文字频道
            for channel in target_category.text_channels:
                # 检查是否被排除
                if not await self._is_channel_excluded(ctx.guild, channel):
                    channels_in_category.append(channel)
        
        if not channels_in_category:
            await ctx.send(f"❌ 分类 `{category_name}` 中没有可总结的频道。")
            return
        
        await ctx.send(f"🔄 开始总结分类 `{category_name}`，共 {len(channels_in_category)} 个频道...")
        
        # 发送到指定频道或当前频道
        summary_channel_id = await self.config.guild(ctx.guild).summary_channel()
        target_channel = ctx.guild.get_channel(summary_channel_id) if summary_channel_id else ctx.channel
        
        # 发送分类标题
        await target_channel.send(f"## 📊 分类总结 - {category_name}\n生成时间: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC")
        
        summaries_data = []  # 收集PDF数据
        
        # 按频道位置排序并总结
        for channel in sorted(channels_in_category, key=lambda c: c.position):
            try:
                summary_embed = await self.generate_channel_summary(channel)
                await target_channel.send(embed=summary_embed)
                
                # 收集PDF数据
                if generate_pdf:
                    summary_text = summary_embed.description or "无总结内容"
                    stats = {}
                    for field in summary_embed.fields:
                        if "消息数量" in field.name:
                            stats['message_count'] = field.value
                        elif "参与人数" in field.name:
                            stats['user_count'] = field.value
                        elif "时间范围" in field.name:
                            stats['time_range'] = field.value
                    
                    summaries_data.append({
                        'category': category_name,
                        'channel_name': channel.name,
                        'summary_text': summary_text,
                        'stats': stats
                    })
                
                log.info(f"成功总结频道 {channel.name} (分类: {category_name}, Guild: {ctx.guild.name})")
                await asyncio.sleep(1)  # 避免速率限制
            except Exception as e:
                log.error(f"总结频道 {channel.name} 时出错 (分类: {category_name}, Guild: {ctx.guild.name}): {e}", exc_info=True)
        
        await ctx.send(f"✅ 分类 `{category_name}` 总结完成！共总结了 {len(channels_in_category)} 个频道。")
        
        # 生成并发送PDF
        if generate_pdf and summaries_data:
            await ctx.send("📄 正在生成PDF报告...")
            report_title = f"{ctx.guild.name} - {category_name} Summary Report"
            pdf_path = await self.generate_pdf_report(ctx.guild, summaries_data, report_title)
            
            if pdf_path and os.path.exists(pdf_path):
                try:
                    await target_channel.send(
                        "📊 总结报告PDF文件：",
                        file=discord.File(pdf_path, filename=f"summary_{category_name}_{datetime.utcnow().strftime('%Y%m%d')}.pdf")
                    )
                    log.info(f"成功发送PDF报告 (分类: {category_name}, Guild: {ctx.guild.name})")
                    # 删除临时文件
                    os.remove(pdf_path)
                except Exception as e:
                    log.error(f"发送PDF文件时出错: {e}", exc_info=True)
                    await ctx.send("❌ PDF文件生成成功但发送失败。")
            else:
                await ctx.send("❌ PDF文件生成失败。请检查日志。")
    
    @summary.group(name="schedule", aliases=["定时", "任务"])
    @checks.admin_or_permissions(manage_guild=True)
    async def schedule(self, ctx: commands.Context):
        """定时任务管理"""
        if ctx.invoked_subcommand is None:
            await ctx.send_help(ctx.command)
    
    @schedule.command(name="add", aliases=["添加", "新增"])
    async def schedule_add(self, ctx: commands.Context, channel: discord.TextChannel, interval_hours: int, run_now: bool = False):
        """添加定时总结任务
        
        参数:
            channel: 要定时总结的频道
            interval_hours: 总结间隔（小时）
            run_now: 是否立即执行一次总结（默认为 False）
        """
        if interval_hours < 1:
            await ctx.send("❌ 间隔时间必须至少为 1 小时。")
            return
        
        async with self.config.guild(ctx.guild).scheduled_tasks() as tasks:
            tasks[str(channel.id)] = {
                "interval": interval_hours,
                "enabled": True,
                "channel_name": channel.name
            }
        
        # 启动定时任务
        self.start_scheduled_task(ctx.guild.id, channel.id, interval_hours)
        
        message = f"✅ 已添加定时任务：每 {interval_hours} 小时总结 {channel.mention}"
        
        # 如果指定立即执行
        if run_now:
            message += "\n🔄 正在立即执行第一次总结..."
            await ctx.send(message)
            async with ctx.typing():
                await self._execute_summary(ctx.guild, channel)
            await ctx.send(f"✅ 首次总结已完成！")
        else:
            await ctx.send(message)
    
    @schedule.command(name="addall", aliases=["添加全部", "新增全部"])
    async def schedule_addall(self, ctx: commands.Context, interval_hours: int, run_now: bool = False):
        """添加定时总结全部频道任务
        
        参数:
            interval_hours: 总结间隔（小时）
            run_now: 是否立即执行一次总结（默认为 False）
        """
        if interval_hours < 1:
            await ctx.send("❌ 间隔时间必须至少为 1 小时。")
            return
        
        async with self.config.guild(ctx.guild).scheduled_tasks() as tasks:
            tasks["0"] = {
                "interval": interval_hours,
                "enabled": True,
                "channel_name": "全部频道",
                "is_all": True
            }
        
        # 启动定时任务（使用 channel_id = 0 表示全服务器）
        self.start_scheduled_task(ctx.guild.id, 0, interval_hours)
        
        message = f"✅ 已添加定时任务：每 {interval_hours} 小时总结全部频道"
        
        # 如果指定立即执行
        if run_now:
            message += "\n🔄 正在立即执行第一次全部频道总结..."
            await ctx.send(message)
            try:
                async with ctx.typing():
                    await self._execute_all_summary(ctx.guild)
            except Exception as e:
                log.warning(f"无法发送 typing 状态: {e}")
                await self._execute_all_summary(ctx.guild)
            await ctx.send(f"✅ 首次全部频道总结已完成！")
        else:
            await ctx.send(message)
    
    @schedule.command(name="remove", aliases=["删除", "移除"])
    async def schedule_remove(self, ctx: commands.Context, channel: discord.TextChannel):
        """移除定时总结任务
        
        参数:
            channel: 要移除任务的频道
        """
        async with self.config.guild(ctx.guild).scheduled_tasks() as tasks:
            if str(channel.id) in tasks:
                del tasks[str(channel.id)]
                
                # 取消任务
                task_key = f"{ctx.guild.id}_{channel.id}"
                if task_key in self.scheduled_jobs:
                    self.scheduled_jobs[task_key].cancel()
                    del self.scheduled_jobs[task_key]
                
                await ctx.send(f"✅ 已移除 {channel.mention} 的定时任务。")
            else:
                await ctx.send(f"❌ 频道 {channel.mention} 没有配置定时任务。")
    
    @schedule.command(name="removeall", aliases=["删除全部", "移除全部"])
    async def schedule_removeall(self, ctx: commands.Context):
        """移除定时总结全部频道任务"""
        async with self.config.guild(ctx.guild).scheduled_tasks() as tasks:
            if "0" in tasks:
                del tasks["0"]
                
                # 取消任务
                task_key = f"{ctx.guild.id}_0"
                if task_key in self.scheduled_jobs:
                    self.scheduled_jobs[task_key].cancel()
                    del self.scheduled_jobs[task_key]
                
                await ctx.send(f"✅ 已移除全部频道的定时任务。")
            else:
                await ctx.send(f"❌ 没有配置全部频道的定时任务。")
    
    @schedule.command(name="list", aliases=["列表", "查看"])
    async def schedule_list(self, ctx: commands.Context):
        """查看所有定时任务"""
        tasks = await self.config.guild(ctx.guild).scheduled_tasks()
        
        if not tasks:
            await ctx.send("📋 当前没有配置任何定时任务。")
            return
        
        embed = discord.Embed(
            title="📋 定时总结任务列表",
            color=discord.Color.blue(),
            timestamp=datetime.utcnow()
        )
        
        for channel_id_str, task_config in tasks.items():
            # 检查是否是全服务器任务
            if channel_id_str == "0":
                channel_name = "🌐 全部频道"
            else:
                channel = ctx.guild.get_channel(int(channel_id_str))
                channel_name = channel.mention if channel else task_config.get("channel_name", "未知频道")
            
            interval = task_config.get("interval", "未知")
            enabled = "✅ 启用" if task_config.get("enabled", False) else "❌ 禁用"
            
            embed.add_field(
                name=f"{channel_name}",
                value=f"间隔: {interval} 小时\n状态: {enabled}",
                inline=True
            )
        
        await ctx.send(embed=embed)
    
    @schedule.command(name="run", aliases=["运行", "执行"])
    async def schedule_run(self, ctx: commands.Context, channel: discord.TextChannel):
        """手动立即执行指定频道的定时总结任务
        
        参数:
            channel: 要执行总结的频道
        """
        tasks = await self.config.guild(ctx.guild).scheduled_tasks()
        
        if str(channel.id) not in tasks:
            await ctx.send(f"❌ 频道 {channel.mention} 没有配置定时任务。")
            return
        
        await ctx.send(f"🔄 正在为 {channel.mention} 生成总结...")
        try:
            async with ctx.typing():
                await self._execute_summary(ctx.guild, channel)
        except Exception as e:
            log.warning(f"无法发送 typing 状态: {e}")
            await self._execute_summary(ctx.guild, channel)
        await ctx.send(f"✅ 总结已完成！")
    
    @schedule.command(name="runall", aliases=["运行全部", "执行全部"])
    async def schedule_runall(self, ctx: commands.Context):
        """手动立即执行全部频道的定时总结任务"""
        tasks = await self.config.guild(ctx.guild).scheduled_tasks()
        
        if "0" not in tasks:
            await ctx.send(f"❌ 没有配置全部频道的定时任务。")
            return
        
        await ctx.send(f"🔄 正在生成全部频道总结，这可能需要一些时间...")
        try:
            async with ctx.typing():
                await self._execute_all_summary(ctx.guild)
        except Exception as e:
            log.warning(f"无法发送 typing 状态: {e}")
            await self._execute_all_summary(ctx.guild)
        await ctx.send(f"✅ 全部频道总结已完成！")
    
    @summary.group(name="config", aliases=["配置", "设置"])
    @checks.admin_or_permissions(manage_guild=True)
    async def config_group(self, ctx: commands.Context):
        """配置管理"""
        if ctx.invoked_subcommand is None:
            await ctx.send_help(ctx.command)
    
    @config_group.command(name="enable", aliases=["启用"])
    async def config_enable(self, ctx: commands.Context):
        """启用聊天总结功能"""
        await self.config.guild(ctx.guild).enabled.set(True)
        await ctx.send("✅ 聊天总结功能已启用。")
    
    @config_group.command(name="disable", aliases=["禁用"])
    async def config_disable(self, ctx: commands.Context):
        """禁用聊天总结功能"""
        await self.config.guild(ctx.guild).enabled.set(False)
        await ctx.send("✅ 聊天总结功能已禁用。")
    
    @config_group.command(name="apikey", aliases=["api"])
    async def config_apikey(self, ctx: commands.Context, api_key: str):
        """设置 OpenAI API Key（私聊发送以保护密钥）
        
        参数:
            api_key: OpenAI API 密钥
        """
        await self.config.guild(ctx.guild).api_key.set(api_key)
        
        try:
            await ctx.message.delete()
        except:
            pass
        
        await ctx.author.send("✅ API Key 已设置成功！")
        await ctx.send("✅ API Key 已配置（已删除你的消息以保护密钥）。")
    
    @config_group.command(name="apibase", aliases=["base"])
    async def config_apibase(self, ctx: commands.Context, api_base: str):
        """设置 API Base URL
        
        参数:
            api_base: API 基础 URL（如：https://api.openai.com/v1）
        """
        await self.config.guild(ctx.guild).api_base.set(api_base)
        await ctx.send(f"✅ API Base URL 已设置为: {api_base}")
    
    @config_group.command(name="model", aliases=["模型"])
    async def config_model(self, ctx: commands.Context, model: str):
        """设置使用的 AI 模型
        
        参数:
            model: 模型名称（如：gpt-3.5-turbo, gpt-4）
        """
        await self.config.guild(ctx.guild).model.set(model)
        await ctx.send(f"✅ AI 模型已设置为: {model}")
    
    @config_group.command(name="maxmessages", aliases=["消息数量"])
    async def config_maxmessages(self, ctx: commands.Context, max_messages: int):
        """设置每次总结的最大消息数量
        
        参数:
            max_messages: 最大消息数量（10-1000）
        """
        if max_messages < 10 or max_messages > 1000:
            await ctx.send("❌ 消息数量必须在 10-1000 之间。")
            return
        
        await self.config.guild(ctx.guild).max_messages.set(max_messages)
        await ctx.send(f"✅ 最大消息数量已设置为: {max_messages}")
    
    @config_group.command(name="summarychannel", aliases=["总结频道"])
    async def config_summarychannel(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None):
        """设置总结结果发送的频道
        
        参数:
            channel: 目标频道（不指定则发送到原频道）
        """
        if channel:
            await self.config.guild(ctx.guild).summary_channel.set(channel.id)
            await ctx.send(f"✅ 总结结果将发送到: {channel.mention}")
        else:
            await self.config.guild(ctx.guild).summary_channel.set(None)
            await ctx.send("✅ 总结结果将发送到原频道。")
    
    @config_group.command(name="exportchannel", aliases=["导出频道"])
    async def config_exportchannel(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None):
        """设置Excel导出文件发送的频道
        
        参数:
            channel: 目标频道（不指定则使用总结频道或当前频道）
        """
        if channel:
            await self.config.guild(ctx.guild).export_channel.set(channel.id)
            await ctx.send(f"✅ Excel导出文件将发送到: {channel.mention}")
        else:
            await self.config.guild(ctx.guild).export_channel.set(None)
            await ctx.send("✅ Excel导出文件将使用总结频道或当前频道。")
    
    @config_group.command(name="exclude", aliases=["排除"])
    async def config_exclude(self, ctx: commands.Context, channel: discord.TextChannel):
        """将频道添加到排除列表（不会被"全部总结"包含）
        
        参数:
            channel: 要排除的频道
        """
        async with self.config.guild(ctx.guild).excluded_channels() as excluded:
            if channel.id not in excluded:
                excluded.append(channel.id)
                await ctx.send(f"✅ 已将 {channel.mention} 添加到排除列表。")
            else:
                await ctx.send(f"❌ {channel.mention} 已在排除列表中。")
    
    @config_group.command(name="include", aliases=["包含"])
    async def config_include(self, ctx: commands.Context, channel: discord.TextChannel):
        """将频道从排除列表中移除
        
        参数:
            channel: 要包含的频道
        """
        async with self.config.guild(ctx.guild).excluded_channels() as excluded:
            if channel.id in excluded:
                excluded.remove(channel.id)
                await ctx.send(f"✅ 已将 {channel.mention} 从排除列表移除。")
            else:
                await ctx.send(f"❌ {channel.mention} 不在排除列表中。")
    
    @config_group.command(name="excludecategory", aliases=["排除分类"])
    async def config_exclude_category(self, ctx: commands.Context, *, category_name: str):
        """将整个分类添加到排除列表（该分类下的所有频道不会被总结）
        
        参数:
            category_name: 要排除的分类名称（使用"未分类"排除没有分类的频道）
        """
        # 检查分类是否存在
        category_exists = False
        if category_name == "未分类":
            category_exists = any(not ch.category for ch in ctx.guild.text_channels)
        else:
            category_exists = any(cat.name == category_name for cat in ctx.guild.categories)
        
        if not category_exists:
            await ctx.send(f"❌ 找不到名为 `{category_name}` 的分类。")
            return
        
        async with self.config.guild(ctx.guild).excluded_categories() as excluded:
            if category_name not in excluded:
                excluded.append(category_name)
                await ctx.send(f"✅ 已将分类 `{category_name}` 添加到排除列表。")
            else:
                await ctx.send(f"❌ 分类 `{category_name}` 已在排除列表中。")
    
    @config_group.command(name="includecategory", aliases=["包含分类"])
    async def config_include_category(self, ctx: commands.Context, *, category_name: str):
        """将分类从排除列表中移除
        
        参数:
            category_name: 要包含的分类名称
        """
        async with self.config.guild(ctx.guild).excluded_categories() as excluded:
            if category_name in excluded:
                excluded.remove(category_name)
                await ctx.send(f"✅ 已将分类 `{category_name}` 从排除列表移除。")
            else:
                await ctx.send(f"❌ 分类 `{category_name}` 不在排除列表中。")
    
    @config_group.command(name="includebots", aliases=["包含机器人"])
    async def config_includebots(self, ctx: commands.Context, include: bool):
        """设置是否包含机器人消息
        
        参数:
            include: True 或 False
        """
        await self.config.guild(ctx.guild).include_bots.set(include)
        status = "包含" if include else "不包含"
        await ctx.send(f"✅ 总结将 {status} 机器人消息。")
    
    @config_group.command(name="exportmaxmessages", aliases=["导出最大消息数"])
    async def config_export_maxmessages(self, ctx: commands.Context, max_messages: int):
        """设置Excel导出的最大消息数量（0表示不限制）
        
        参数:
            max_messages: 最大消息数量
        """
        if max_messages < 0:
            await ctx.send("❌ 消息数量不能为负数。")
            return
        
        await self.config.guild(ctx.guild).export_max_messages.set(max_messages)
        if max_messages == 0:
            await ctx.send(f"✅ 导出最大消息数量已设置为: 不限制")
        else:
            await ctx.send(f"✅ 导出最大消息数量已设置为: {max_messages}")
    
    @config_group.command(name="exportexclude", aliases=["导出排除"])
    async def config_export_exclude(self, ctx: commands.Context, channel: discord.TextChannel):
        """将频道添加到导出排除列表（不会被"全部导出"包含）
        
        参数:
            channel: 要排除的频道
        """
        async with self.config.guild(ctx.guild).export_excluded_channels() as excluded:
            if channel.id not in excluded:
                excluded.append(channel.id)
                await ctx.send(f"✅ 已将 {channel.mention} 添加到导出排除列表。")
            else:
                await ctx.send(f"❌ {channel.mention} 已在导出排除列表中。")
    
    @config_group.command(name="exportinclude", aliases=["导出包含"])
    async def config_export_include(self, ctx: commands.Context, channel: discord.TextChannel):
        """将频道从导出排除列表中移除
        
        参数:
            channel: 要包含的频道
        """
        async with self.config.guild(ctx.guild).export_excluded_channels() as excluded:
            if channel.id in excluded:
                excluded.remove(channel.id)
                await ctx.send(f"✅ 已将 {channel.mention} 从导出排除列表移除。")
            else:
                await ctx.send(f"❌ {channel.mention} 不在导出排除列表中。")
    
    @config_group.command(name="exportexcludecategory", aliases=["导出排除分类"])
    async def config_export_exclude_category(self, ctx: commands.Context, *, category_name: str):
        """将分类添加到导出排除列表（该分类下的所有频道都不会被导出）
        
        参数:
            category_name: 要排除的分类名称
        """
        async with self.config.guild(ctx.guild).export_excluded_categories() as excluded:
            if category_name not in excluded:
                excluded.append(category_name)
                await ctx.send(f"✅ 已将分类 `{category_name}` 添加到导出排除列表。")
            else:
                await ctx.send(f"❌ 分类 `{category_name}` 已在导出排除列表中。")
    
    @config_group.command(name="exportincludecategory", aliases=["导出包含分类"])
    async def config_export_include_category(self, ctx: commands.Context, *, category_name: str):
        """将分类从导出排除列表中移除
        
        参数:
            category_name: 要包含的分类名称
        """
        async with self.config.guild(ctx.guild).export_excluded_categories() as excluded:
            if category_name in excluded:
                excluded.remove(category_name)
                await ctx.send(f"✅ 已将分类 `{category_name}` 从导出排除列表移除。")
            else:
                await ctx.send(f"❌ 分类 `{category_name}` 不在导出排除列表中。")
    
    @config_group.command(name="show", aliases=["显示", "查看"])
    async def config_show(self, ctx: commands.Context):
        """显示当前配置"""
        config = await self.config.guild(ctx.guild).all()
        
        api_key_status = "✅ 已配置" if config["api_key"] else "❌ 未配置"
        enabled_status = "✅ 已启用" if config["enabled"] else "❌ 已禁用"
        
        summary_channel = ctx.guild.get_channel(config["summary_channel"]) if config["summary_channel"] else None
        summary_channel_text = summary_channel.mention if summary_channel else "原频道"
        
        export_channel = ctx.guild.get_channel(config["export_channel"]) if config["export_channel"] else None
        export_channel_text = export_channel.mention if export_channel else "总结频道/当前频道"
        
        excluded_channels = [
            ctx.guild.get_channel(ch_id).mention 
            for ch_id in config["excluded_channels"] 
            if ctx.guild.get_channel(ch_id)
        ]
        excluded_channels_text = ", ".join(excluded_channels) if excluded_channels else "无"
        
        excluded_categories = config.get("excluded_categories", [])
        excluded_categories_text = ", ".join([f"`{cat}`" for cat in excluded_categories]) if excluded_categories else "无"
        
        export_excluded_channels = [
            ctx.guild.get_channel(ch_id).mention 
            for ch_id in config.get("export_excluded_channels", [])
            if ctx.guild.get_channel(ch_id)
        ]
        export_excluded_channels_text = ", ".join(export_excluded_channels) if export_excluded_channels else "无"
        
        export_excluded_categories = config.get("export_excluded_categories", [])
        export_excluded_categories_text = ", ".join([f"`{cat}`" for cat in export_excluded_categories]) if export_excluded_categories else "无"
        
        embed = discord.Embed(
            title="⚙️ 聊天总结配置",
            color=discord.Color.blue(),
            timestamp=datetime.utcnow()
        )
        
        embed.add_field(name="功能状态", value=enabled_status, inline=True)
        embed.add_field(name="API Key", value=api_key_status, inline=True)
        embed.add_field(name="AI 模型", value=config["model"], inline=True)
        embed.add_field(name="API Base", value=config["api_base"], inline=False)
        embed.add_field(name="总结最大消息数", value=str(config["max_messages"]), inline=True)
        embed.add_field(name="导出最大消息数", value=str(config.get("export_max_messages", 1000)), inline=True)
        embed.add_field(name="包含机器人", value="是" if config["include_bots"] else "否", inline=True)
        embed.add_field(name="总结发送频道", value=summary_channel_text, inline=True)
        embed.add_field(name="导出发送频道", value=export_channel_text, inline=True)
        embed.add_field(name="总结排除频道", value=excluded_channels_text, inline=False)
        embed.add_field(name="总结排除分类", value=excluded_categories_text, inline=False)
        embed.add_field(name="导出排除频道", value=export_excluded_channels_text, inline=False)
        embed.add_field(name="导出排除分类", value=export_excluded_categories_text, inline=False)
        embed.add_field(name="总结定时任务数", value=str(len(config["scheduled_tasks"])), inline=True)
        embed.add_field(name="导出定时任务数", value=str(len(config.get("export_tasks", {}))), inline=True)
        
        await ctx.send(embed=embed)
    
    @config_group.command(name="testfont", aliases=["测试字体"])
    async def config_testfont(self, ctx: commands.Context):
        """测试系统中文字体可用性"""
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # 测试字体列表
        chinese_fonts = [
            # macOS
            ('/System/Library/Fonts/PingFang.ttc', 0, 'macOS PingFang'),
            ('/System/Library/Fonts/STHeiti Light.ttc', 0, 'macOS STHeiti'),
            ('/System/Library/Fonts/Hiragino Sans GB.ttc', 0, 'macOS Hiragino'),
            # Linux
            ('/usr/share/fonts/truetype/arphic/uming.ttc', 0, 'Linux AR PL UMing'),
            ('/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf', None, 'Linux Droid Sans'),
            ('/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc', 0, 'Linux WenQuanYi'),
            # Windows
            ('C:\\Windows\\Fonts\\msyh.ttc', 0, 'Windows 微软雅黑'),
            ('C:\\Windows\\Fonts\\simhei.ttf', None, 'Windows 黑体'),
            ('C:\\Windows\\Fonts\\simsun.ttc', 0, 'Windows 宋体'),
        ]
        
        embed = discord.Embed(
            title="🔤 中文字体检测",
            description="检测系统中可用的PDF中文字体",
            color=discord.Color.blue(),
            timestamp=datetime.utcnow()
        )
        
        found_fonts = []
        available_fonts = []
        
        for font_path, subfont_index, font_name in chinese_fonts:
            if os.path.exists(font_path):
                found_fonts.append(f"✅ {font_name}\n路径: `{font_path}`")
                
                # 尝试注册测试
                try:
                    test_name = f"Test_{len(available_fonts)}"
                    if subfont_index is not None:
                        pdfmetrics.registerFont(TTFont(test_name, font_path, subfontIndex=subfont_index))
                    else:
                        pdfmetrics.registerFont(TTFont(test_name, font_path))
                    available_fonts.append(font_name)
                except Exception as e:
                    found_fonts[-1] += f"\n⚠️ 注册失败: {str(e)[:50]}"
        
        if found_fonts:
            embed.add_field(
                name="找到的字体",
                value="\n\n".join(found_fonts),
                inline=False
            )
        else:
            embed.add_field(
                name="找到的字体",
                value="❌ 未找到任何中文字体",
                inline=False
            )
        
        if available_fonts:
            embed.add_field(
                name="✅ 可用于PDF",
                value=", ".join(available_fonts),
                inline=False
            )
            embed.color = discord.Color.green()
        else:
            embed.add_field(
                name="❌ PDF生成问题",
                value="未找到可用的中文字体，PDF中的中文将显示为方块",
                inline=False
            )
            embed.color = discord.Color.red()
        
        embed.add_field(
            name="💡 解决方法",
            value="如果没有找到字体，请安装中文字体包：\n"
                  "• macOS: 已内置中文字体\n"
                  "• Linux: `sudo apt-get install fonts-arphic-uming`\n"
                  "• Windows: 确保安装了微软雅黑字体",
            inline=False
        )
        
        await ctx.send(embed=embed)
    
    async def generate_excel_report(self, channel: discord.TextChannel, max_messages: int = None) -> str:
        """生成单个频道的Excel报告
        
        参数:
            channel: Discord频道
            max_messages: 最大消息数量（None表示不限制）
        
        返回:
            Excel文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            guild = channel.guild
            category_name = channel.category.name if channel.category else "未分类"
            
            # 如果没有指定最大消息数，使用配置的值
            if max_messages is None:
                max_messages = await self.config.guild(guild).max_messages()
            
            include_bots = await self.config.guild(guild).include_bots()
            
            log.info(f"开始生成Excel报告 (频道: {channel.name}, 最大消息数: {max_messages})")
            
            # 获取消息
            messages = []
            async for message in channel.history(limit=max_messages if max_messages > 0 else None):
                if not include_bots and message.author.bot:
                    continue
                messages.append(message)
            
            messages.reverse()  # 按时间顺序排列
            
            log.info(f"获取到 {len(messages)} 条消息")
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "聊天记录"
            
            # 设置标题样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 定义列标题
            headers = [
                "消息ID", "时间", "用户名", "用户ID", "用户昵称", 
                "消息内容", "Embed内容", "附件", "回复消息ID", "反应", 
                "是否编辑", "编辑时间", "是否置顶", "提及用户"
            ]
            
            # 写入标题行
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入数据
            for row_num, message in enumerate(messages, 2):
                # 消息ID
                ws.cell(row=row_num, column=1, value=str(message.id))
                
                # 时间
                ws.cell(row=row_num, column=2, value=message.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                
                # 用户名
                ws.cell(row=row_num, column=3, value=str(message.author))
                
                # 用户ID
                ws.cell(row=row_num, column=4, value=str(message.author.id))
                
                # 用户昵称（服务器内昵称）
                nickname = message.author.display_name if hasattr(message.author, 'display_name') else str(message.author)
                ws.cell(row=row_num, column=5, value=nickname)
                
                # 消息内容
                content = message.content[:32767] if message.content else ""  # Excel单元格最大长度
                ws.cell(row=row_num, column=6, value=content)
                
                # Embed内容
                embed_content = ""
                if message.embeds:
                    embed_parts = []
                    for embed in message.embeds:
                        embed_info = []
                        if embed.title:
                            embed_info.append(f"标题: {embed.title}")
                        if embed.description:
                            embed_info.append(f"描述: {embed.description[:500]}")  # 限制长度
                        if embed.url:
                            embed_info.append(f"链接: {embed.url}")
                        if embed.author and embed.author.name:
                            embed_info.append(f"作者: {embed.author.name}")
                        if embed.fields:
                            for field in embed.fields:
                                embed_info.append(f"{field.name}: {field.value[:200]}")
                        if embed.footer and embed.footer.text:
                            embed_info.append(f"页脚: {embed.footer.text}")
                        if embed.image:
                            embed_info.append(f"图片: {embed.image.url}")
                        if embed.thumbnail:
                            embed_info.append(f"缩略图: {embed.thumbnail.url}")
                        if embed.video:
                            embed_info.append(f"视频: {embed.video.url}")
                        
                        if embed_info:
                            embed_parts.append(" | ".join(embed_info))
                    
                    embed_content = "\n---\n".join(embed_parts)
                    # 确保不超过Excel单元格限制
                    if len(embed_content) > 32767:
                        embed_content = embed_content[:32764] + "..."
                
                ws.cell(row=row_num, column=7, value=embed_content)
                
                # 附件
                attachments = ", ".join([att.url for att in message.attachments]) if message.attachments else ""
                ws.cell(row=row_num, column=8, value=attachments)
                
                # 回复消息ID
                reply_id = str(message.reference.message_id) if message.reference else ""
                ws.cell(row=row_num, column=9, value=reply_id)
                
                # 反应
                reactions = ", ".join([f"{reaction.emoji}({reaction.count})" for reaction in message.reactions]) if message.reactions else ""
                ws.cell(row=row_num, column=10, value=reactions)
                
                # 是否编辑
                ws.cell(row=row_num, column=11, value="是" if message.edited_at else "否")
                
                # 编辑时间
                edited_time = message.edited_at.strftime('%Y-%m-%d %H:%M:%S') if message.edited_at else ""
                ws.cell(row=row_num, column=12, value=edited_time)
                
                # 是否置顶
                ws.cell(row=row_num, column=13, value="是" if message.pinned else "否")
                
                # 提及用户
                mentions = ", ".join([str(user) for user in message.mentions]) if message.mentions else ""
                ws.cell(row=row_num, column=14, value=mentions)
            
            # 调整列宽
            column_widths = {
                1: 20,  # 消息ID
                2: 20,  # 时间
                3: 20,  # 用户名
                4: 20,  # 用户ID
                5: 20,  # 用户昵称
                6: 50,  # 消息内容
                7: 50,  # Embed内容
                8: 40,  # 附件
                9: 20,  # 回复消息ID
                10: 30, # 反应
                11: 12, # 是否编辑
                12: 20, # 编辑时间
                13: 12, # 是否置顶
                14: 30, # 提及用户
            }
            
            for col_num, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            # 冻结首行
            ws.freeze_panes = "A2"
            
            # 添加统计信息工作表
            stats_ws = wb.create_sheet("统计信息")
            stats_ws['A1'] = "统计项目"
            stats_ws['B1'] = "数值"
            stats_ws['A1'].font = header_font
            stats_ws['B1'].font = header_font
            
            stats_data = [
                ("频道名称", f"{category_name} / {channel.name}"),
                ("消息总数", len(messages)),
                ("参与用户数", len(set(m.author.id for m in messages))),
                ("时间范围", f"{messages[0].created_at.strftime('%Y-%m-%d %H:%M:%S')} - {messages[-1].created_at.strftime('%Y-%m-%d %H:%M:%S')}" if messages else "无"),
                ("包含附件的消息", sum(1 for m in messages if m.attachments)),
                ("包含Embed的消息", sum(1 for m in messages if m.embeds)),
                ("编辑过的消息", sum(1 for m in messages if m.edited_at)),
                ("置顶消息", sum(1 for m in messages if m.pinned)),
                ("回复消息", sum(1 for m in messages if m.reference)),
                ("生成时间", datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S') + " UTC"),
            ]
            
            for idx, (stat_name, stat_value) in enumerate(stats_data, 2):
                stats_ws[f'A{idx}'] = stat_name
                stats_ws[f'B{idx}'] = stat_value
            
            stats_ws.column_dimensions['A'].width = 20
            stats_ws.column_dimensions['B'].width = 50
            
            # 保存文件
            temp_dir = tempfile.gettempdir()
            filename = f"{category_name}-{channel.name}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            # 清理文件名中的非法字符
            filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.', '（', '）', '(', ')'))
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            log.info(f"Excel报告已生成: {filepath}")
            
            return filepath
            
        except ImportError:
            log.error("openpyxl库未安装，无法生成Excel报告")
            return None
        except Exception as e:
            log.error(f"生成Excel报告时出错: {e}", exc_info=True)
            return None
    
    async def generate_multi_channel_excel_report(self, guild: discord.Guild, channels_dict: dict, report_title: str, max_messages: int = None) -> str:
        """生成多频道合并的Excel报告
        
        参数:
            guild: Discord服务器
            channels_dict: 频道字典 {category_name: [channels]}
            report_title: 报告标题
            max_messages: 每个频道的最大消息数量
        
        返回:
            Excel文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            log.info(f"开始生成多频道Excel报告 (服务器: {guild.name}, 分类数: {len(channels_dict)})")
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 设置样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 定义列标题
            headers = [
                "消息ID", "时间", "用户名", "用户ID", "用户昵称", 
                "消息内容", "Embed内容", "附件", "回复消息ID", "反应", 
                "是否编辑", "编辑时间", "是否置顶", "提及用户"
            ]
            
            # 如果没有指定最大消息数，使用配置的值
            if max_messages is None:
                max_messages = await self.config.guild(guild).max_messages()
            
            include_bots = await self.config.guild(guild).include_bots()
            
            # 统计信息
            total_messages = 0
            total_channels = 0
            all_stats = []
            
            # 按分类名称排序
            sorted_categories = sorted(channels_dict.keys(), key=lambda x: (x == "未分类", x))
            
            # 为每个频道创建工作表
            for category_name in sorted_categories:
                channels = channels_dict[category_name]
                
                for channel in sorted(channels, key=lambda c: c.position):
                    try:
                        log.info(f"正在处理频道: {category_name} / {channel.name}")
                        
                        # 获取消息
                        messages = []
                        async for message in channel.history(limit=max_messages if max_messages > 0 else None):
                            if not include_bots and message.author.bot:
                                continue
                            messages.append(message)
                        
                        messages.reverse()  # 按时间顺序排列
                        
                        if not messages:
                            log.info(f"频道 {channel.name} 没有消息，跳过")
                            continue
                        
                        total_messages += len(messages)
                        total_channels += 1
                        
                        # 创建工作表（限制工作表名称长度）
                        sheet_name = f"{category_name}-{channel.name}"
                        # Excel工作表名称限制为31个字符
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:28] + "..."
                        # 移除非法字符
                        sheet_name = "".join(c for c in sheet_name if c not in [':', '\\', '/', '?', '*', '[', ']'])
                        
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # 写入标题行
                        for col_num, header in enumerate(headers, 1):
                            cell = ws.cell(row=1, column=col_num)
                            cell.value = header
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        
                        # 写入数据
                        for row_num, message in enumerate(messages, 2):
                            # 消息ID
                            ws.cell(row=row_num, column=1, value=str(message.id))
                            
                            # 时间
                            ws.cell(row=row_num, column=2, value=message.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                            
                            # 用户名
                            ws.cell(row=row_num, column=3, value=str(message.author))
                            
                            # 用户ID
                            ws.cell(row=row_num, column=4, value=str(message.author.id))
                            
                            # 用户昵称
                            nickname = message.author.display_name if hasattr(message.author, 'display_name') else str(message.author)
                            ws.cell(row=row_num, column=5, value=nickname)
                            
                            # 消息内容
                            content = message.content[:32767] if message.content else ""
                            ws.cell(row=row_num, column=6, value=content)
                            
                            # Embed内容
                            embed_content = ""
                            if message.embeds:
                                embed_parts = []
                                for embed in message.embeds:
                                    embed_info = []
                                    if embed.title:
                                        embed_info.append(f"标题: {embed.title}")
                                    if embed.description:
                                        embed_info.append(f"描述: {embed.description[:500]}")
                                    if embed.url:
                                        embed_info.append(f"链接: {embed.url}")
                                    if embed.author and embed.author.name:
                                        embed_info.append(f"作者: {embed.author.name}")
                                    if embed.fields:
                                        for field in embed.fields:
                                            embed_info.append(f"{field.name}: {field.value[:200]}")
                                    if embed.footer and embed.footer.text:
                                        embed_info.append(f"页脚: {embed.footer.text}")
                                    if embed.image:
                                        embed_info.append(f"图片: {embed.image.url}")
                                    if embed.thumbnail:
                                        embed_info.append(f"缩略图: {embed.thumbnail.url}")
                                    if embed.video:
                                        embed_info.append(f"视频: {embed.video.url}")
                                    
                                    if embed_info:
                                        embed_parts.append(" | ".join(embed_info))
                                
                                embed_content = "\n---\n".join(embed_parts)
                                if len(embed_content) > 32767:
                                    embed_content = embed_content[:32764] + "..."
                            
                            ws.cell(row=row_num, column=7, value=embed_content)
                            
                            # 附件
                            attachments = ", ".join([att.url for att in message.attachments]) if message.attachments else ""
                            ws.cell(row=row_num, column=8, value=attachments)
                            
                            # 回复消息ID
                            reply_id = str(message.reference.message_id) if message.reference else ""
                            ws.cell(row=row_num, column=9, value=reply_id)
                            
                            # 反应
                            reactions = ", ".join([f"{reaction.emoji}({reaction.count})" for reaction in message.reactions]) if message.reactions else ""
                            ws.cell(row=row_num, column=10, value=reactions)
                            
                            # 是否编辑
                            ws.cell(row=row_num, column=11, value="是" if message.edited_at else "否")
                            
                            # 编辑时间
                            edited_time = message.edited_at.strftime('%Y-%m-%d %H:%M:%S') if message.edited_at else ""
                            ws.cell(row=row_num, column=12, value=edited_time)
                            
                            # 是否置顶
                            ws.cell(row=row_num, column=13, value="是" if message.pinned else "否")
                            
                            # 提及用户
                            mentions = ", ".join([str(user) for user in message.mentions]) if message.mentions else ""
                            ws.cell(row=row_num, column=14, value=mentions)
                        
                        # 调整列宽
                        column_widths = {
                            1: 20, 2: 20, 3: 20, 4: 20, 5: 20,
                            6: 50, 7: 50, 8: 40, 9: 20, 10: 30,
                            11: 12, 12: 20, 13: 12, 14: 30,
                        }
                        
                        for col_num, width in column_widths.items():
                            ws.column_dimensions[get_column_letter(col_num)].width = width
                        
                        # 冻结首行
                        ws.freeze_panes = "A2"
                        
                        # 收集统计信息
                        all_stats.append({
                            'category': category_name,
                            'channel': channel.name,
                            'messages': len(messages),
                            'users': len(set(m.author.id for m in messages)),
                            'time_range': f"{messages[0].created_at.strftime('%Y-%m-%d %H:%M:%S')} - {messages[-1].created_at.strftime('%Y-%m-%d %H:%M:%S')}",
                            'attachments': sum(1 for m in messages if m.attachments),
                            'embeds': sum(1 for m in messages if m.embeds),
                            'edited': sum(1 for m in messages if m.edited_at),
                            'pinned': sum(1 for m in messages if m.pinned),
                            'replies': sum(1 for m in messages if m.reference),
                        })
                        
                        log.info(f"已添加工作表: {sheet_name} ({len(messages)} 条消息)")
                        
                    except Exception as e:
                        log.error(f"处理频道 {channel.name} 时出错: {e}", exc_info=True)
                        continue
            
            # 创建汇总统计工作表（放在最前面）
            summary_ws = wb.create_sheet(title="📊 汇总统计", index=0)
            summary_ws['A1'] = "服务器名称"
            summary_ws['B1'] = guild.name
            summary_ws['A2'] = "报告标题"
            summary_ws['B2'] = report_title
            summary_ws['A3'] = "生成时间"
            summary_ws['B3'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S') + " UTC"
            summary_ws['A4'] = "总频道数"
            summary_ws['B4'] = total_channels
            summary_ws['A5'] = "总消息数"
            summary_ws['B5'] = total_messages
            
            # 设置标题样式
            for row in range(1, 6):
                summary_ws[f'A{row}'].font = header_font
            
            # 添加详细统计表头
            summary_ws['A7'] = "分类"
            summary_ws['B7'] = "频道"
            summary_ws['C7'] = "消息数"
            summary_ws['D7'] = "用户数"
            summary_ws['E7'] = "附件"
            summary_ws['F7'] = "Embed"
            summary_ws['G7'] = "编辑"
            summary_ws['H7'] = "置顶"
            summary_ws['I7'] = "回复"
            summary_ws['J7'] = "时间范围"
            
            for col in range(1, 11):
                cell = summary_ws.cell(row=7, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入各频道统计
            for idx, stat in enumerate(all_stats, 8):
                summary_ws[f'A{idx}'] = stat['category']
                summary_ws[f'B{idx}'] = stat['channel']
                summary_ws[f'C{idx}'] = stat['messages']
                summary_ws[f'D{idx}'] = stat['users']
                summary_ws[f'E{idx}'] = stat['attachments']
                summary_ws[f'F{idx}'] = stat['embeds']
                summary_ws[f'G{idx}'] = stat['edited']
                summary_ws[f'H{idx}'] = stat['pinned']
                summary_ws[f'I{idx}'] = stat['replies']
                summary_ws[f'J{idx}'] = stat['time_range']
            
            # 调整列宽
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 20
            summary_ws.column_dimensions['C'].width = 12
            summary_ws.column_dimensions['D'].width = 12
            summary_ws.column_dimensions['E'].width = 10
            summary_ws.column_dimensions['F'].width = 10
            summary_ws.column_dimensions['G'].width = 10
            summary_ws.column_dimensions['H'].width = 10
            summary_ws.column_dimensions['I'].width = 10
            summary_ws.column_dimensions['J'].width = 40
            
            # 保存文件
            temp_dir = tempfile.gettempdir()
            filename = f"{guild.name}_{report_title}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.', '（', '）', '(', ')'))
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            log.info(f"多频道Excel报告已生成: {filepath} (包含 {total_channels} 个频道，{total_messages} 条消息)")
            
            return filepath
            
        except ImportError:
            log.error("openpyxl库未安装，无法生成Excel报告")
            return None
        except Exception as e:
            log.error(f"生成多频道Excel报告时出错: {e}", exc_info=True)
            return None
    
    @summary.group(name="export", aliases=["导出"])
    @commands.guild_only()
    async def export_group(self, ctx: commands.Context):
        """导出聊天记录到Excel"""
        if ctx.invoked_subcommand is None:
            await ctx.send_help(ctx.command)
    
    @export_group.command(name="channel", aliases=["频道"])
    async def export_channel(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None, max_messages: int = 0):
        """导出指定频道的聊天记录到Excel
        
        参数:
            channel: 要导出的频道（不指定则导出当前频道）
            max_messages: 最大消息数量（0表示使用配置的默认值）
        """
        if not await self.config.guild(ctx.guild).enabled():
            await ctx.send("❌ 聊天总结功能未启用。请管理员使用 `[p]summary enable` 启用。")
            return
        
        target_channel = channel or ctx.channel
        
        await ctx.send(f"📊 正在导出频道 {target_channel.mention} 的聊天记录，请稍候...")
        
        try:
            async with ctx.typing():
                # 如果未指定，使用配置的导出最大消息数
                if max_messages == 0:
                    default_max = await self.config.guild(ctx.guild).export_max_messages()
                    max_msgs = default_max if default_max > 0 else None
                else:
                    max_msgs = max_messages
                excel_path = await self.generate_excel_report(target_channel, max_msgs)
                
                if excel_path and os.path.exists(excel_path):
                    category_name = target_channel.category.name if target_channel.category else "未分类"
                    await send_channel.send(
                        f"✅ Excel报告已生成！",
                        file=discord.File(excel_path, filename=f"{category_name}-{target_channel.name}.xlsx")
                    )
                    # 删除临时文件
                    os.remove(excel_path)
                    log.info(f"成功发送Excel报告 (频道: {target_channel.name})")
                else:
                    await ctx.send("❌ Excel报告生成失败。请确保已安装 openpyxl 库。")
        except Exception as e:
            log.error(f"导出Excel时出错: {e}", exc_info=True)
            await ctx.send(f"❌ 导出失败: {str(e)}")
    
    @export_group.command(name="all", aliases=["全部"])
    @checks.admin_or_permissions(manage_guild=True)
    async def export_all(self, ctx: commands.Context, max_messages: int = 0, single_file: bool = True):
        """导出所有频道的聊天记录到Excel
        
        参数:
            max_messages: 每个频道的最大消息数量（0表示使用配置的默认值）
            single_file: 是否合并到单个文件（True=单文件，False=每个频道一个文件）
        """
        if not await self.config.guild(ctx.guild).enabled():
            await ctx.send("❌ 聊天总结功能未启用。")
            return
        
        file_mode = "单文件模式" if single_file else "多文件模式"
        await ctx.send(f"🔄 开始导出所有频道的聊天记录（{file_mode}），这可能需要较长时间...")
        
        # 按分类分组频道
        categories_dict = defaultdict(list)
        
        for channel in ctx.guild.text_channels:
            if await self._is_channel_excluded_from_export(ctx.guild, channel):
                continue
            
            category_name = channel.category.name if channel.category else "未分类"
            categories_dict[category_name].append(channel)
        
        if not categories_dict:
            await ctx.send("❌ 没有可导出的频道。")
            return
        
        # 发送到指定频道或当前频道（优先使用导出频道）
        export_channel_id = await self.config.guild(ctx.guild).export_channel()
        summary_channel_id = await self.config.guild(ctx.guild).summary_channel()
        
        if export_channel_id:
            target_channel = ctx.guild.get_channel(export_channel_id)
        elif summary_channel_id:
            target_channel = ctx.guild.get_channel(summary_channel_id)
        else:
            target_channel = ctx.channel
        
        if not target_channel:
            target_channel = ctx.channel
        
        # 如果未指定，使用配置的导出最大消息数
        if max_messages == 0:
            default_max = await self.config.guild(ctx.guild).export_max_messages()
            max_msgs = default_max if default_max > 0 else None
        else:
            max_msgs = max_messages
        
        if single_file:
            # 单文件模式：所有频道合并到一个Excel文件
            try:
                await target_channel.send("📄 正在生成合并Excel报告...")
                report_title = "全服务器聊天记录"
                excel_path = await self.generate_multi_channel_excel_report(
                    ctx.guild, categories_dict, report_title, max_msgs
                )
                
                if excel_path and os.path.exists(excel_path):
                    file_size = os.path.getsize(excel_path) / (1024 * 1024)  # MB
                    await target_channel.send(
                        f"✅ Excel报告已生成（文件大小: {file_size:.2f} MB）",
                        file=discord.File(excel_path, filename=f"{ctx.guild.name}_全服务器聊天记录.xlsx")
                    )
                    os.remove(excel_path)
                    log.info(f"成功发送合并Excel报告")
                else:
                    await ctx.send("❌ Excel报告生成失败。请检查日志。")
            except Exception as e:
                log.error(f"生成合并Excel报告时出错: {e}", exc_info=True)
                await ctx.send(f"❌ 导出失败: {str(e)}")
        else:
            # 多文件模式：每个频道一个文件
            total_channels = 0
            failed_channels = []
            
            # 按分类名称排序
            sorted_categories = sorted(categories_dict.keys(), key=lambda x: (x == "未分类", x))
            
            for category_name in sorted_categories:
                channels = categories_dict[category_name]
                
                await target_channel.send(f"\n## 📁 正在导出分类: {category_name}\n")
                
                for channel in sorted(channels, key=lambda c: c.position):
                    try:
                        excel_path = await self.generate_excel_report(channel, max_msgs)
                        
                        if excel_path and os.path.exists(excel_path):
                            await target_channel.send(
                                f"📊 {category_name} / {channel.name}",
                                file=discord.File(excel_path, filename=f"{category_name}-{channel.name}.xlsx")
                            )
                            os.remove(excel_path)
                            total_channels += 1
                            log.info(f"成功导出频道 {channel.name}")
                        else:
                            failed_channels.append(f"{category_name}/{channel.name}")
                            log.error(f"导出频道 {channel.name} 失败")
                        
                        await asyncio.sleep(2)  # 避免速率限制
                    except Exception as e:
                        failed_channels.append(f"{category_name}/{channel.name}")
                        log.error(f"导出频道 {channel.name} 时出错: {e}", exc_info=True)
            
            # 发送完成消息
            if failed_channels:
                await ctx.send(f"⚠️ 导出完成！成功导出 {total_channels} 个频道，{len(failed_channels)} 个频道失败。\n失败的频道: {', '.join(failed_channels)}")
            else:
                await ctx.send(f"✅ 导出完成！共成功导出 {total_channels} 个频道。")
    
    @export_group.command(name="category", aliases=["分类"])
    @checks.admin_or_permissions(manage_guild=True)
    async def export_category(self, ctx: commands.Context, category_name: str, max_messages: int = 0, single_file: bool = True):
        """导出指定分类下所有频道的聊天记录到Excel
        
        参数:
            category_name: 分类名称（使用"未分类"导出没有分类的频道）
            max_messages: 每个频道的最大消息数量（0表示使用配置的默认值）
            single_file: 是否合并到单个文件（True=单文件，False=每个频道一个文件）
        """
        if not await self.config.guild(ctx.guild).enabled():
            await ctx.send("❌ 聊天总结功能未启用。")
            return
        
        # 查找分类下的频道
        channels_in_category = []
        
        if category_name == "未分类":
            for channel in ctx.guild.text_channels:
                if not channel.category:
                    if not await self._is_channel_excluded_from_export(ctx.guild, channel):
                        channels_in_category.append(channel)
        else:
            target_category = None
            for category in ctx.guild.categories:
                if category.name == category_name:
                    target_category = category
                    break
            
            if not target_category:
                await ctx.send(f"❌ 找不到名为 `{category_name}` 的分类。")
                return
            
            for channel in target_category.text_channels:
                if not await self._is_channel_excluded_from_export(ctx.guild, channel):
                    channels_in_category.append(channel)
        
        if not channels_in_category:
            await ctx.send(f"❌ 分类 `{category_name}` 中没有可导出的频道。")
            return
        
        file_mode = "单文件模式" if single_file else "多文件模式"
        await ctx.send(f"🔄 开始导出分类 `{category_name}`，共 {len(channels_in_category)} 个频道（{file_mode}）...")
        
        # 发送到指定频道或当前频道（优先使用导出频道）
        export_channel_id = await self.config.guild(ctx.guild).export_channel()
        summary_channel_id = await self.config.guild(ctx.guild).summary_channel()
        
        if export_channel_id:
            target_channel = ctx.guild.get_channel(export_channel_id)
        elif summary_channel_id:
            target_channel = ctx.guild.get_channel(summary_channel_id)
        else:
            target_channel = ctx.channel
        
        if not target_channel:
            target_channel = ctx.channel
        
        # 如果未指定，使用配置的导出最大消息数
        if max_messages == 0:
            default_max = await self.config.guild(ctx.guild).export_max_messages()
            max_msgs = default_max if default_max > 0 else None
        else:
            max_msgs = max_messages
        
        if single_file:
            # 单文件模式：所有频道合并到一个Excel文件
            try:
                await target_channel.send(f"📄 正在生成分类 `{category_name}` 的合并Excel报告...")
                
                # 构建频道字典
                categories_dict = {category_name: channels_in_category}
                report_title = f"{category_name}_聊天记录"
                
                excel_path = await self.generate_multi_channel_excel_report(
                    ctx.guild, categories_dict, report_title, max_msgs
                )
                
                if excel_path and os.path.exists(excel_path):
                    file_size = os.path.getsize(excel_path) / (1024 * 1024)  # MB
                    await target_channel.send(
                        f"✅ 分类 `{category_name}` 的Excel报告已生成（文件大小: {file_size:.2f} MB）",
                        file=discord.File(excel_path, filename=f"{category_name}_聊天记录.xlsx")
                    )
                    os.remove(excel_path)
                    log.info(f"成功发送分类 {category_name} 的合并Excel报告")
                else:
                    await ctx.send("❌ Excel报告生成失败。请检查日志。")
            except Exception as e:
                log.error(f"生成分类合并Excel报告时出错: {e}", exc_info=True)
                await ctx.send(f"❌ 导出失败: {str(e)}")
        else:
            # 多文件模式：每个频道一个文件
            total_channels = 0
            failed_channels = []
            
            for channel in sorted(channels_in_category, key=lambda c: c.position):
                try:
                    excel_path = await self.generate_excel_report(channel, max_msgs)
                    
                    if excel_path and os.path.exists(excel_path):
                        await target_channel.send(
                            f"📊 {category_name} / {channel.name}",
                            file=discord.File(excel_path, filename=f"{category_name}-{channel.name}.xlsx")
                        )
                        os.remove(excel_path)
                        total_channels += 1
                        log.info(f"成功导出频道 {channel.name}")
                    else:
                        failed_channels.append(channel.name)
                        log.error(f"导出频道 {channel.name} 失败")
                    
                    await asyncio.sleep(2)  # 避免速率限制
                except Exception as e:
                    failed_channels.append(channel.name)
                    log.error(f"导出频道 {channel.name} 时出错: {e}", exc_info=True)
            
            # 发送完成消息
            if failed_channels:
                await ctx.send(f"⚠️ 导出完成！成功导出 {total_channels} 个频道，{len(failed_channels)} 个频道失败。\n失败的频道: {', '.join(failed_channels)}")
            else:
                await ctx.send(f"✅ 分类 `{category_name}` 导出完成！共成功导出 {total_channels} 个频道。")
    
    @export_group.group(name="schedule", aliases=["定时", "任务"])
    @checks.admin_or_permissions(manage_guild=True)
    async def export_schedule(self, ctx: commands.Context):
        """Excel导出定时任务管理"""
        if ctx.invoked_subcommand is None:
            await ctx.send_help(ctx.command)
    
    @export_schedule.command(name="addall", aliases=["添加全部"])
    async def export_schedule_addall(self, ctx: commands.Context, interval_hours: int, single_file: bool = True, max_messages: int = 0, run_now: bool = False):
        """添加定时导出所有频道的任务
        
        参数:
            interval_hours: 导出间隔（小时）
            single_file: 是否合并到单个文件（默认True）
            max_messages: 每个频道的最大消息数量（0=使用默认值）
            run_now: 是否立即执行一次（默认False）
        """
        if interval_hours < 1:
            await ctx.send("❌ 间隔时间必须至少为 1 小时。")
            return
        
        task_id = "export_all"
        task_config = {
            "type": "all",
            "target": "",
            "interval": interval_hours,
            "enabled": True,
            "single_file": single_file,
            "max_messages": max_messages
        }
        
        async with self.config.guild(ctx.guild).export_tasks() as tasks:
            tasks[task_id] = task_config
        
        # 启动定时任务
        self.start_export_task(ctx.guild.id, task_id, task_config)
        
        file_mode = "单文件" if single_file else "多文件"
        message = f"✅ 已添加定时导出任务：每 {interval_hours} 小时导出所有频道（{file_mode}模式）"
        
        if run_now:
            message += "\n🔄 正在立即执行第一次导出..."
            await ctx.send(message)
            await self._execute_export_task(ctx.guild, task_config)
            await ctx.send(f"✅ 首次导出已完成！")
        else:
            await ctx.send(message)
    
    @export_schedule.command(name="addcategory", aliases=["添加分类"])
    async def export_schedule_addcategory(self, ctx: commands.Context, category_name: str, interval_hours: int, single_file: bool = True, max_messages: int = 0, run_now: bool = False):
        """添加定时导出指定分类的任务
        
        参数:
            category_name: 分类名称
            interval_hours: 导出间隔（小时）
            single_file: 是否合并到单个文件（默认True）
            max_messages: 每个频道的最大消息数量（0=使用默认值）
            run_now: 是否立即执行一次（默认False）
        """
        if interval_hours < 1:
            await ctx.send("❌ 间隔时间必须至少为 1 小时。")
            return
        
        task_id = f"export_cat_{category_name}"
        task_config = {
            "type": "category",
            "target": category_name,
            "interval": interval_hours,
            "enabled": True,
            "single_file": single_file,
            "max_messages": max_messages
        }
        
        async with self.config.guild(ctx.guild).export_tasks() as tasks:
            tasks[task_id] = task_config
        
        # 启动定时任务
        self.start_export_task(ctx.guild.id, task_id, task_config)
        
        file_mode = "单文件" if single_file else "多文件"
        message = f"✅ 已添加定时导出任务：每 {interval_hours} 小时导出分类 `{category_name}`（{file_mode}模式）"
        
        if run_now:
            message += "\n🔄 正在立即执行第一次导出..."
            await ctx.send(message)
            await self._execute_export_task(ctx.guild, task_config)
            await ctx.send(f"✅ 首次导出已完成！")
        else:
            await ctx.send(message)
    
    @export_schedule.command(name="addchannel", aliases=["添加频道"])
    async def export_schedule_addchannel(self, ctx: commands.Context, channel: discord.TextChannel, interval_hours: int, max_messages: int = 0, run_now: bool = False):
        """添加定时导出指定频道的任务
        
        参数:
            channel: 要导出的频道
            interval_hours: 导出间隔（小时）
            max_messages: 最大消息数量（0=使用默认值）
            run_now: 是否立即执行一次（默认False）
        """
        if interval_hours < 1:
            await ctx.send("❌ 间隔时间必须至少为 1 小时。")
            return
        
        task_id = f"export_ch_{channel.id}"
        task_config = {
            "type": "channel",
            "target": str(channel.id),
            "interval": interval_hours,
            "enabled": True,
            "single_file": False,  # 单频道不需要合并
            "max_messages": max_messages
        }
        
        async with self.config.guild(ctx.guild).export_tasks() as tasks:
            tasks[task_id] = task_config
        
        # 启动定时任务
        self.start_export_task(ctx.guild.id, task_id, task_config)
        
        message = f"✅ 已添加定时导出任务：每 {interval_hours} 小时导出 {channel.mention}"
        
        if run_now:
            message += "\n🔄 正在立即执行第一次导出..."
            await ctx.send(message)
            await self._execute_export_task(ctx.guild, task_config)
            await ctx.send(f"✅ 首次导出已完成！")
        else:
            await ctx.send(message)
    
    @export_schedule.command(name="remove", aliases=["删除", "移除"])
    async def export_schedule_remove(self, ctx: commands.Context, task_id: str):
        """移除导出定时任务
        
        参数:
            task_id: 任务ID（使用 list 命令查看）
        """
        async with self.config.guild(ctx.guild).export_tasks() as tasks:
            if task_id in tasks:
                del tasks[task_id]
                
                # 取消任务
                task_key = f"{ctx.guild.id}_{task_id}"
                if task_key in self.export_jobs:
                    self.export_jobs[task_key].cancel()
                    del self.export_jobs[task_key]
                
                await ctx.send(f"✅ 已移除导出定时任务: {task_id}")
            else:
                await ctx.send(f"❌ 找不到任务ID: {task_id}")
    
    @export_schedule.command(name="list", aliases=["列表", "查看"])
    async def export_schedule_list(self, ctx: commands.Context):
        """查看所有导出定时任务"""
        tasks = await self.config.guild(ctx.guild).export_tasks()
        
        if not tasks:
            await ctx.send("📋 当前没有配置任何导出定时任务。")
            return
        
        embed = discord.Embed(
            title="📋 导出定时任务列表",
            color=discord.Color.blue(),
            timestamp=datetime.utcnow()
        )
        
        for task_id, task_config in tasks.items():
            task_type = task_config.get("type", "unknown")
            target = task_config.get("target", "")
            interval = task_config.get("interval", "未知")
            enabled = "✅ 启用" if task_config.get("enabled", False) else "❌ 禁用"
            single_file = task_config.get("single_file", True)
            max_messages = task_config.get("max_messages", 0)
            
            if task_type == "all":
                task_name = "🌐 所有频道"
            elif task_type == "category":
                task_name = f"📁 分类: {target}"
            elif task_type == "channel":
                channel = ctx.guild.get_channel(int(target))
                task_name = f"💬 频道: {channel.mention if channel else target}"
            else:
                task_name = f"❓ {task_type}"
            
            file_mode = "单文件" if single_file else "多文件"
            max_msg_text = f"{max_messages}条" if max_messages > 0 else "默认"
            
            embed.add_field(
                name=f"{task_name} (ID: {task_id})",
                value=f"间隔: {interval} 小时\n模式: {file_mode}\n消息数: {max_msg_text}\n状态: {enabled}",
                inline=True
            )
        
        await ctx.send(embed=embed)
    
    @export_schedule.command(name="run", aliases=["运行", "执行"])
    async def export_schedule_run(self, ctx: commands.Context, task_id: str):
        """手动立即执行指定的导出定时任务
        
        参数:
            task_id: 任务ID（使用 list 命令查看）
        """
        tasks = await self.config.guild(ctx.guild).export_tasks()
        
        if task_id not in tasks:
            await ctx.send(f"❌ 找不到任务ID: {task_id}")
            return
        
        task_config = tasks[task_id]
        await ctx.send(f"🔄 正在执行导出任务: {task_id}...")
        
        try:
            await self._execute_export_task(ctx.guild, task_config)
        except Exception as e:
            log.error(f"手动执行导出任务失败: {e}", exc_info=True)
            await ctx.send(f"❌ 执行失败: {str(e)}")

